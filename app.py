"""
Zoom Participation Grader

Minimal package requirements:
- streamlit
- pandas
- openpyxl
- rapidfuzz
"""

from __future__ import annotations

import base64
import csv
import hashlib
import html
import io
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process


# =========================
# Constants and defaults
# =========================

APP_TITLE = "Zoom Participation Grader"
UNKNOWN_SPEAKER = "Unknown Speaker"
REVIEWER_OVERRIDE_OPTIONS = ["", "Award", "Do_Not_Award", "Manual_Review"]
BONUS_POLICY_OPTIONS = [
    "attended only",
    "spoke only",
    "attended and spoke",
    "weighted score",
]

TRANSCRIPT_COLUMNS = [
    "source_file",
    "meeting_name",
    "meeting_id",
    "meeting_key",
    "timestamp_start",
    "timestamp_end",
    "raw_speaker",
    "utterance_text",
    "word_count",
    "speaker_normalized",
    "speaker_is_unknown",
]

SPEAKER_AGG_COLUMNS = [
    "meeting_key",
    "meeting_name",
    "meeting_id",
    "raw_speaker",
    "canonical_speaker_candidate",
    "total_turns",
    "total_words",
    "first_speaking_time",
    "last_speaking_time",
    "speaking_span_minutes",
    "unique_speaking_intervals",
]

ATTENDANCE_COLUMNS = [
    "source_file",
    "meeting_name",
    "meeting_id",
    "meeting_key",
    "participant_name",
    "email",
    "join_time",
    "leave_time",
    "duration_minutes",
    "participant_normalized",
    "row_valid",
]

ATTENDANCE_AGG_COLUMNS = [
    "meeting_key",
    "meeting_name",
    "meeting_id",
    "participant_name",
    "email",
    "total_duration_minutes",
    "first_join",
    "last_leave",
    "join_count",
]

ROSTER_COLUMNS = [
    "roster_name",
    "roster_email",
    "roster_name_normalized",
    "roster_row_id",
]

ALIAS_COLUMNS = [
    "alias_name",
    "canonical_name",
    "alias_normalized",
    "canonical_normalized",
]

MATCHED_COLUMNS = [
    "meeting_key",
    "meeting_name",
    "meeting_id",
    "evidence_source",
    "source_file",
    "canonical_student_name",
    "raw_name",
    "raw_attendance_name",
    "raw_transcript_speaker",
    "email",
    "attendance_minutes",
    "join_count",
    "first_join",
    "last_leave",
    "speaking_turns",
    "speaking_words",
    "first_speaking_time",
    "last_speaking_time",
    "speaking_span_minutes",
    "match_method",
    "match_confidence",
    "matched_from_source",
    "ambiguous_match",
    "identity_review_flag",
    "review_reason",
    "raw_name_generic",
    "speaker_is_unknown",
    "student_merge_key",
]

FINAL_COLUMNS = [
    "meeting_name",
    "meeting_id",
    "canonical_student_name",
    "raw_attendance_name",
    "raw_transcript_speaker",
    "email",
    "attended",
    "spoke",
    "attendance_minutes",
    "speaking_turns",
    "speaking_words",
    "speaking_span_minutes",
    "join_count",
    "match_method",
    "match_confidence",
    "manual_review",
    "recommend_award",
    "reviewer_notes",
    "reviewer_override",
    "decision_reason",
]

MAX_PREVIEW_ROWS = 200

HEADER_SYNONYMS = {
    "meeting_name": [
        "topic",
        "meeting topic",
        "meeting name",
        "session name",
        "class name",
    ],
    "meeting_id": ["meeting id", "meetingid", "id", "session id"],
    "participant_name": [
        "name",
        "name (original name)",
        "participant",
        "participant name",
        "user name",
        "display name",
        "attendee name",
        "participants",
    ],
    "email": [
        "user email",
        "email",
        "email address",
        "participant email",
        "attendee email",
    ],
    "join_time": ["join time", "join_time", "joined", "join", "time joined"],
    "leave_time": ["leave time", "leave_time", "left", "leave", "time left"],
    "duration_minutes": [
        "duration",
        "duration (minutes)",
        "duration minutes",
        "minutes in meeting",
        "time in meeting",
    ],
}

ROSTER_HEADER_SYNONYMS = {
    "name": [
        "name",
        "student name",
        "full name",
        "preferred name",
        "display name",
    ],
    "email": [
        "email",
        "student email",
        "email address",
        "school email",
        "primary email",
    ],
}

ALIAS_HEADER_SYNONYMS = {
    "alias_name": ["alias", "alias name", "speaker", "raw name", "display name"],
    "canonical_name": [
        "canonical",
        "canonical name",
        "student name",
        "roster name",
        "official name",
    ],
}

GENERIC_DEVICE_PATTERNS = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in [
        r"^\s*$",
        r"^unknown speaker$",
        r"^iphone$",
        r"^ipad$",
        r"^galaxy$",
        r"^android$",
        r"^phone$",
        r"^guest$",
        r"^guest user$",
        r"^zoom user$",
        r"^user$",
        r"^participant$",
        r"^student$",
        r"^mobile$",
        r"^tablet$",
        r"^lenovo$",
        r"^samsung$",
        r"^device$",
    ]
]

DEVICE_SUFFIX_PATTERNS = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in [
        r"\b(?:iphone|ipad|galaxy|android|phone|mobile|tablet)\b",
        r"\((?:guest|iphone|ipad|galaxy|android|phone|mobile|tablet|zoom user)\)",
        r"\b(?:guest|zoom user)\b",
    ]
]

TIMESTAMP_RANGE_RE = re.compile(
    r"(?P<start>\d{1,2}:\d{2}:\d{2}(?:[.,]\d{1,3})?)\s*-->\s*"
    r"(?P<end>\d{1,2}:\d{2}:\d{2}(?:[.,]\d{1,3})?)"
)

DEFAULT_ATTENDANCE_THRESHOLD = 30
DEFAULT_WORD_THRESHOLD = 20
DEFAULT_TURN_THRESHOLD = 2
DEFAULT_SPAN_THRESHOLD = 5
DEFAULT_FUZZY_THRESHOLD = 0.86
DEFAULT_SAFE_AUTO_THRESHOLD = 0.93
DEFAULT_ALIAS_CONFIDENCE = 0.93
DEFAULT_AMBIGUITY_MARGIN = 0.03

DEFAULT_ATTENDANCE_WEIGHT = 0.5
DEFAULT_WORD_WEIGHT = 0.3
DEFAULT_TURN_WEIGHT = 0.2
DEFAULT_ATTENDANCE_CAP = 60
DEFAULT_WORD_CAP = 120
DEFAULT_TURN_CAP = 8
DEFAULT_WEIGHTED_THRESHOLD = 0.7

CONFIG_SHEET_NAME = "Config"
RAW_TRANSCRIPT_SHEET = "Raw_Transcript"
RAW_ATTENDANCE_SHEET = "Raw_Attendance"
AGG_SPEAKERS_SHEET = "Aggregated_Speakers"
AGG_ATTENDANCE_SHEET = "Aggregated_Attendance"
MATCHED_SHEET = "Matched_Students"
AWARD_SHEET = "Award"
DO_NOT_AWARD_SHEET = "Do_Not_Award"
MANUAL_REVIEW_SHEET = "Manual_Review"
AUDIT_SHEET = "Audit_Log"


# =========================
# Dataclasses
# =========================


@dataclass
class AppConfig:
    attendance_threshold_minutes: float = DEFAULT_ATTENDANCE_THRESHOLD
    words_threshold: int = DEFAULT_WORD_THRESHOLD
    turns_threshold: int = DEFAULT_TURN_THRESHOLD
    span_threshold_enabled: bool = False
    span_threshold_minutes: float = DEFAULT_SPAN_THRESHOLD
    fuzzy_threshold: float = DEFAULT_FUZZY_THRESHOLD
    safe_auto_approval_threshold: float = DEFAULT_SAFE_AUTO_THRESHOLD
    alias_confidence: float = DEFAULT_ALIAS_CONFIDENCE
    ambiguity_margin: float = DEFAULT_AMBIGUITY_MARGIN
    bonus_policy_mode: str = "attended and spoke"
    attendance_weight: float = DEFAULT_ATTENDANCE_WEIGHT
    word_weight: float = DEFAULT_WORD_WEIGHT
    turn_weight: float = DEFAULT_TURN_WEIGHT
    attendance_cap_minutes: float = DEFAULT_ATTENDANCE_CAP
    word_cap: int = DEFAULT_WORD_CAP
    turn_cap: int = DEFAULT_TURN_CAP
    weighted_threshold: float = DEFAULT_WEIGHTED_THRESHOLD
    combine_all_meetings: bool = True
    selected_meetings: tuple[str, ...] = ()
    search_text: str = ""
    show_only_award: bool = False
    show_only_do_not_award: bool = False
    show_only_manual_review: bool = False
    show_only_unmatched: bool = False
    show_only_low_confidence: bool = False


@dataclass
class ValidationIssue:
    level: str
    source_type: str
    source_file: str
    message: str
    row_hint: str = ""


@dataclass
class MatchOutcome:
    raw_name: str
    canonical_name: str
    match_method: str
    match_confidence: float
    matched_from_source: str
    ambiguous: bool
    review_reason: str = ""


# =========================
# Utility helpers
# =========================


def empty_df(columns: Sequence[str]) -> pd.DataFrame:
    """Return an empty DataFrame with the requested columns."""
    return pd.DataFrame(columns=list(columns))


def sha1_hex(data: bytes) -> str:
    """Return the SHA-1 hex digest for file deduplication."""
    return hashlib.sha1(data).hexdigest()


def safe_decode(data: bytes) -> str:
    """Decode uploaded bytes with a forgiving UTF strategy."""
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def normalize_header(value: Any) -> str:
    """Normalize a column header for matching."""
    text = stringify(value).lower()
    text = re.sub(r"[\s_/\\-]+", " ", text)
    text = re.sub(r"[^a-z0-9 ]+", "", text)
    return re.sub(r"\s+", " ", text).strip()


def stringify(value: Any) -> str:
    """Convert a value into a safe display string."""
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_email(value: Any) -> str:
    """Normalize an email for exact comparisons."""
    return stringify(value).lower()


def collapse_whitespace(text: str) -> str:
    """Collapse repeated whitespace."""
    return re.sub(r"\s+", " ", text).strip()


def strip_html_tags(text: str) -> str:
    """Remove VTT/HTML-like tags from transcript text."""
    cleaned = re.sub(r"<[^>]+>", " ", stringify(text))
    return collapse_whitespace(html.unescape(cleaned))


def remove_device_noise(text: str) -> str:
    """Remove common device and guest suffix noise from names."""
    cleaned = stringify(text)
    for pattern in DEVICE_SUFFIX_PATTERNS:
        cleaned = pattern.sub(" ", cleaned)
    cleaned = re.sub(r"[^\w\s@.]+", " ", cleaned)
    return collapse_whitespace(cleaned)


def looks_generic_device_name(text: str) -> bool:
    """Return True when a display name is too generic to trust automatically."""
    raw = collapse_whitespace(remove_device_noise(text).lower())
    if not raw:
        return True
    return any(pattern.match(raw) for pattern in GENERIC_DEVICE_PATTERNS)


def normalize_name(text: Any) -> str:
    """Normalize a name for exact and fuzzy matching."""
    raw = remove_device_noise(stringify(text)).lower()
    raw = re.sub(r"[^\w\s@.]+", " ", raw)
    raw = re.sub(r"\b(the|mr|mrs|ms|dr)\b", " ", raw)
    raw = re.sub(r"\s+", " ", raw).strip()
    return raw


def join_unique(values: Iterable[Any], sep: str = " | ") -> str:
    """Join unique non-empty values while preserving first-seen order."""
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        text = stringify(value)
        if not text:
            continue
        if text not in seen:
            seen.add(text)
            ordered.append(text)
    return sep.join(ordered)


def first_non_blank(values: Iterable[Any], default: str = "") -> str:
    """Return the first non-empty string from a sequence."""
    for value in values:
        text = stringify(value)
        if text:
            return text
    return default


def parse_datetime_value(value: Any) -> pd.Timestamp:
    """Parse a date or datetime string into a pandas timestamp."""
    text = stringify(value)
    if not text:
        return pd.NaT
    return pd.to_datetime(text, errors="coerce")


def parse_duration_minutes(
    value: Any, join_time: pd.Timestamp, leave_time: pd.Timestamp
) -> Optional[float]:
    """Parse duration minutes, falling back to join/leave timestamps."""
    text = stringify(value)
    if text:
        plain = text.replace("minutes", "").replace("minute", "").strip()
        if re.fullmatch(r"\d+(?:\.\d+)?", plain):
            return round(float(plain), 2)
        if re.fullmatch(r"\d{1,2}:\d{2}:\d{2}", plain):
            hours, minutes, seconds = [int(part) for part in plain.split(":")]
            return round(hours * 60 + minutes + seconds / 60, 2)
        if re.fullmatch(r"\d{1,2}:\d{2}", plain):
            minutes, seconds = [int(part) for part in plain.split(":")]
            return round(minutes + seconds / 60, 2)
    if pd.notna(join_time) and pd.notna(leave_time):
        delta = (leave_time - join_time).total_seconds() / 60
        return round(max(delta, 0), 2)
    return None


def extract_meeting_id(text: str) -> str:
    """Extract a likely Zoom meeting ID from text or file names."""
    match = re.search(r"(?<!\d)(\d{9,11})(?!\d)", stringify(text))
    return match.group(1) if match else ""


def guess_meeting_name(file_name: str) -> str:
    """Guess a clean meeting name from a file name."""
    stem = Path(file_name).stem
    stem = re.sub(r"[_-]+", " ", stem)
    stem = re.sub(r"(?<!\d)\d{9,11}(?!\d)", " ", stem)
    stem = re.sub(r"\b(audio transcript|transcript|participants?|attendance)\b", "", stem, flags=re.I)
    stem = re.sub(r"\s+", " ", stem).strip(" -_")
    return stem or Path(file_name).stem


def build_meeting_key(meeting_name: str, meeting_id: str, source_file: str) -> str:
    """Build a stable meeting key for joins across files."""
    normalized_name = normalize_name(meeting_name) or normalize_name(guess_meeting_name(source_file))
    safe_id = stringify(meeting_id)
    return f"{safe_id or 'noid'}::{normalized_name or normalize_name(source_file)}"


def parse_vtt_timestamp(timestamp_text: str) -> Optional[float]:
    """Parse a VTT timestamp to seconds."""
    text = stringify(timestamp_text).replace(",", ".")
    try:
        hours, minutes, seconds = text.split(":")
        return (
            int(hours) * 3600
            + int(minutes) * 60
            + float(seconds)
        )
    except (TypeError, ValueError):
        return None


def seconds_to_timestamp(seconds: Optional[float]) -> str:
    """Convert seconds to an HH:MM:SS.mmm transcript timestamp."""
    if seconds is None or pd.isna(seconds):
        return ""
    seconds = max(float(seconds), 0.0)
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = seconds - hours * 3600 - minutes * 60
    return f"{hours:02d}:{minutes:02d}:{secs:06.3f}"


def detect_column_candidates(
    columns: Sequence[str], synonyms: Sequence[str]
) -> list[str]:
    """Return best-effort matching columns for a semantic field."""
    normalized_synonyms = {normalize_header(item) for item in synonyms}
    scored: list[tuple[int, str]] = []
    for column in columns:
        header = normalize_header(column)
        score = 0
        if header in normalized_synonyms:
            score = 3
        elif any(header in synonym or synonym in header for synonym in normalized_synonyms):
            score = 2
        elif any(
            token and token in set(header.split())
            for synonym in normalized_synonyms
            for token in synonym.split()
        ):
            score = 1
        if score > 0:
            scored.append((score, column))
    scored.sort(key=lambda item: (-item[0], list(columns).index(item[1])))
    return [column for _, column in scored]


def detect_columns(
    columns: Sequence[str], synonym_map: dict[str, Sequence[str]]
) -> dict[str, str]:
    """Detect canonical fields from uploaded table columns."""
    detected: dict[str, str] = {}
    for canonical, synonyms in synonym_map.items():
        candidates = detect_column_candidates(columns, synonyms)
        if candidates:
            detected[canonical] = candidates[0]
    return detected


def make_unique_headers(headers: Sequence[Any]) -> list[str]:
    """Ensure duplicate CSV headers remain addressable."""
    counts: dict[str, int] = defaultdict(int)
    unique_headers: list[str] = []
    for header in headers:
        text = stringify(header) or "Unnamed"
        counts[text] += 1
        unique_headers.append(text if counts[text] == 1 else f"{text}_{counts[text]}")
    return unique_headers


def read_csv_with_detected_header(
    file_name: str, data: bytes, expected_headers: dict[str, Sequence[str]]
) -> tuple[pd.DataFrame, list[ValidationIssue]]:
    """Read a CSV while tolerating preambles and malformed rows."""
    issues: list[ValidationIssue] = []
    text = safe_decode(data)
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader]
    if not rows:
        issues.append(
            ValidationIssue("error", "csv", file_name, "Uploaded CSV is empty.")
        )
        return empty_df([]), issues

    expected = {
        normalize_header(name)
        for synonym_list in expected_headers.values()
        for name in synonym_list
    }
    best_idx = 0
    best_score = -1
    for idx, row in enumerate(rows[:15]):
        normalized = [normalize_header(cell) for cell in row]
        score = sum(1 for cell in normalized if cell in expected)
        if score > best_score and len(row) >= 2:
            best_score = score
            best_idx = idx

    headers = make_unique_headers(rows[best_idx])
    width = len(headers)
    normalized_rows: list[list[str]] = []
    for raw_row in rows[best_idx + 1 :]:
        if not any(stringify(cell) for cell in raw_row):
            continue
        row = list(raw_row[:width]) + [""] * max(width - len(raw_row), 0)
        normalized_rows.append(row[:width])

    if not normalized_rows:
        issues.append(
            ValidationIssue(
                "warning",
                "csv",
                file_name,
                "CSV file contains headers but no data rows.",
            )
        )
    return pd.DataFrame(normalized_rows, columns=headers), issues


def read_table_file(
    file_name: str, data: bytes, expected_headers: dict[str, Sequence[str]]
) -> tuple[pd.DataFrame, list[ValidationIssue]]:
    """Read CSV/XLSX data into a DataFrame."""
    suffix = Path(file_name).suffix.lower()
    if suffix == ".xlsx":
        try:
            frame = pd.read_excel(io.BytesIO(data), dtype=str).fillna("")
            return frame, []
        except Exception as exc:  # pragma: no cover - defensive error path
            return empty_df([]), [
                ValidationIssue(
                    "error",
                    "xlsx",
                    file_name,
                    f"Could not read Excel file: {exc}",
                )
            ]
    return read_csv_with_detected_header(file_name, data, expected_headers)


def uploaded_to_memory(
    uploaded_files: Sequence[Any], source_type: str
) -> tuple[list[dict[str, Any]], list[ValidationIssue]]:
    """Read Streamlit uploads once, deduplicate them, and keep bytes in memory."""
    records: list[dict[str, Any]] = []
    issues: list[ValidationIssue] = []
    seen_hashes: set[str] = set()
    for file in uploaded_files:
        data = file.getvalue()
        if not data:
            issues.append(
                ValidationIssue(
                    "warning",
                    source_type,
                    file.name,
                    "Empty file skipped.",
                )
            )
            continue
        digest = sha1_hex(data)
        if digest in seen_hashes:
            issues.append(
                ValidationIssue(
                    "warning",
                    source_type,
                    file.name,
                    "Duplicate upload ignored because the file content matches another upload.",
                )
            )
            continue
        seen_hashes.add(digest)
        records.append(
            {"name": file.name, "data": data, "sha1": digest, "source_type": source_type}
        )
    return records, issues


def choose_common_value(series: pd.Series) -> str:
    """Choose the most common non-blank string, falling back to the first non-blank value."""
    values = [stringify(item) for item in series if stringify(item)]
    if not values:
        return ""
    most_common = Counter(values).most_common(1)
    return most_common[0][0] if most_common else values[0]


def safe_min_datetime(series: pd.Series) -> pd.Timestamp:
    """Return the minimum valid timestamp or NaT."""
    valid = pd.to_datetime(series, errors="coerce")
    return valid.min() if valid.notna().any() else pd.NaT


def safe_max_datetime(series: pd.Series) -> pd.Timestamp:
    """Return the maximum valid timestamp or NaT."""
    valid = pd.to_datetime(series, errors="coerce")
    return valid.max() if valid.notna().any() else pd.NaT


# =========================
# Transcript parser
# =========================


def parse_transcript_payload(lines: Sequence[str]) -> tuple[str, str]:
    """Extract a speaker label and utterance text from a VTT payload block."""
    speaker = ""
    text_parts: list[str] = []
    for line in lines:
        raw_line = html.unescape(stringify(line))
        if not raw_line:
            continue
        tagged = re.match(r"^<v(?:\.[^>\s]+)?\s+([^>]+)>(.*)$", raw_line, flags=re.I)
        if tagged:
            if not speaker:
                speaker = strip_html_tags(tagged.group(1))
            cleaned_text = strip_html_tags(tagged.group(2))
            if cleaned_text:
                text_parts.append(cleaned_text)
            continue

        cleaned = strip_html_tags(raw_line)
        if not speaker:
            colon_match = re.match(r"^([^:]{1,80}):\s+(.+)$", cleaned)
            if colon_match:
                candidate = collapse_whitespace(colon_match.group(1))
                if len(candidate.split()) <= 8:
                    speaker = candidate
                    cleaned = collapse_whitespace(colon_match.group(2))
        if cleaned:
            text_parts.append(cleaned)

    utterance = collapse_whitespace(" ".join(text_parts))
    return speaker or UNKNOWN_SPEAKER, utterance


def parse_transcript_file(
    file_name: str, data: bytes
) -> tuple[pd.DataFrame, list[ValidationIssue]]:
    """Parse a Zoom VTT transcript into structured rows."""
    issues: list[ValidationIssue] = []
    text = safe_decode(data)
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    meeting_name = guess_meeting_name(file_name)
    meeting_id = extract_meeting_id(file_name) or extract_meeting_id(text[:1000])
    meeting_key = build_meeting_key(meeting_name, meeting_id, file_name)

    records: list[dict[str, Any]] = []
    idx = 0
    while idx < len(lines):
        current = lines[idx].lstrip("\ufeff").strip()
        if not current or current.upper() == "WEBVTT":
            idx += 1
            continue
        if current.startswith("NOTE") or current.startswith("Kind:") or current.startswith("Language:"):
            idx += 1
            continue
        if current.isdigit() and idx + 1 < len(lines) and "-->" in lines[idx + 1]:
            idx += 1
            current = lines[idx].strip()
        if "-->" not in current:
            idx += 1
            continue

        match = TIMESTAMP_RANGE_RE.search(current)
        if not match:
            issues.append(
                ValidationIssue(
                    "warning",
                    "transcript",
                    file_name,
                    "Malformed timestamp block skipped.",
                    row_hint=f"Line {idx + 1}",
                )
            )
            idx += 1
            continue

        start_seconds = parse_vtt_timestamp(match.group("start"))
        end_seconds = parse_vtt_timestamp(match.group("end"))
        idx += 1
        payload: list[str] = []
        while idx < len(lines) and lines[idx].strip():
            payload.append(lines[idx])
            idx += 1

        if not payload:
            issues.append(
                ValidationIssue(
                    "warning",
                    "transcript",
                    file_name,
                    "Transcript cue had no utterance text.",
                    row_hint=f"Timestamp {match.group('start')}",
                )
            )
            idx += 1
            continue

        raw_speaker, utterance = parse_transcript_payload(payload)
        speaker_normalized = normalize_name(raw_speaker)
        speaker_is_unknown = raw_speaker == UNKNOWN_SPEAKER or speaker_normalized in {
            "",
            normalize_name(UNKNOWN_SPEAKER),
        }
        word_count = len(re.findall(r"\b[\w']+\b", utterance))

        records.append(
            {
                "source_file": file_name,
                "meeting_name": meeting_name,
                "meeting_id": meeting_id,
                "meeting_key": meeting_key,
                "timestamp_start": seconds_to_timestamp(start_seconds),
                "timestamp_end": seconds_to_timestamp(end_seconds),
                "raw_speaker": raw_speaker or UNKNOWN_SPEAKER,
                "utterance_text": utterance,
                "word_count": word_count,
                "speaker_normalized": speaker_normalized,
                "speaker_is_unknown": bool(speaker_is_unknown),
                "start_seconds": start_seconds,
                "end_seconds": end_seconds,
                "interval_key": f"{seconds_to_timestamp(start_seconds)}->{seconds_to_timestamp(end_seconds)}",
            }
        )
        idx += 1

    if not records:
        issues.append(
            ValidationIssue(
                "warning",
                "transcript",
                file_name,
                "No transcript cues could be parsed from this VTT file.",
            )
        )
        return empty_df(TRANSCRIPT_COLUMNS + ["start_seconds", "end_seconds", "interval_key"]), issues

    return pd.DataFrame(records), issues


def aggregate_transcript_speakers(raw_transcript_df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate transcript rows to one speaker per meeting."""
    if raw_transcript_df.empty:
        return empty_df(SPEAKER_AGG_COLUMNS + ["first_speaking_seconds", "last_speaking_seconds", "speaker_is_unknown", "source_file"])

    grouped = (
        raw_transcript_df.groupby(
            ["meeting_key", "meeting_name", "meeting_id", "raw_speaker"],
            dropna=False,
            as_index=False,
        )
        .agg(
            source_file=("source_file", join_unique),
            total_turns=("utterance_text", "count"),
            total_words=("word_count", "sum"),
            first_speaking_seconds=("start_seconds", "min"),
            last_speaking_seconds=("end_seconds", "max"),
            unique_speaking_intervals=("interval_key", "nunique"),
            speaker_is_unknown=("speaker_is_unknown", "all"),
        )
    )
    grouped["first_speaking_time"] = grouped["first_speaking_seconds"].apply(seconds_to_timestamp)
    grouped["last_speaking_time"] = grouped["last_speaking_seconds"].apply(seconds_to_timestamp)
    grouped["speaking_span_minutes"] = (
        (grouped["last_speaking_seconds"] - grouped["first_speaking_seconds"]).clip(lower=0) / 60
    ).round(2)
    grouped["canonical_speaker_candidate"] = grouped["raw_speaker"]
    return grouped


# =========================
# Attendance parser
# =========================


def parse_attendance_file(
    file_name: str, data: bytes
) -> tuple[pd.DataFrame, list[ValidationIssue]]:
    """Parse a Zoom participant CSV into normalized attendance rows."""
    issues: list[ValidationIssue] = []
    table, table_issues = read_csv_with_detected_header(file_name, data, HEADER_SYNONYMS)
    issues.extend(table_issues)
    if table.empty:
        return empty_df(ATTENDANCE_COLUMNS), issues

    detected = detect_columns(table.columns, HEADER_SYNONYMS)
    if "participant_name" not in detected and "email" not in detected:
        issues.append(
            ValidationIssue(
                "error",
                "attendance",
                file_name,
                "Could not find a participant name or email column in the attendance CSV.",
            )
        )
        return empty_df(ATTENDANCE_COLUMNS), issues

    rows: list[dict[str, Any]] = []
    default_meeting_name = guess_meeting_name(file_name)
    default_meeting_id = extract_meeting_id(file_name)

    for row_number, row in table.iterrows():
        participant_name = stringify(row.get(detected.get("participant_name", ""), ""))
        email = normalize_email(row.get(detected.get("email", ""), ""))
        if not participant_name and email:
            participant_name = email

        if not participant_name and not email:
            issues.append(
                ValidationIssue(
                    "warning",
                    "attendance",
                    file_name,
                    "Row skipped because both participant name and email are missing.",
                    row_hint=f"Row {row_number + 2}",
                )
            )
            continue

        join_time = parse_datetime_value(row.get(detected.get("join_time", ""), ""))
        leave_time = parse_datetime_value(row.get(detected.get("leave_time", ""), ""))
        duration = parse_duration_minutes(
            row.get(detected.get("duration_minutes", ""), ""),
            join_time,
            leave_time,
        )
        if duration is None:
            issues.append(
                ValidationIssue(
                    "warning",
                    "attendance",
                    file_name,
                    "Duration could not be parsed; row kept for review.",
                    row_hint=f"Row {row_number + 2}",
                )
            )

        meeting_name = stringify(row.get(detected.get("meeting_name", ""), "")) or default_meeting_name
        meeting_id = stringify(row.get(detected.get("meeting_id", ""), "")) or default_meeting_id
        meeting_key = build_meeting_key(meeting_name, meeting_id, file_name)

        rows.append(
            {
                "source_file": file_name,
                "meeting_name": meeting_name,
                "meeting_id": meeting_id,
                "meeting_key": meeting_key,
                "participant_name": participant_name,
                "email": email,
                "join_time": join_time,
                "leave_time": leave_time,
                "duration_minutes": duration,
                "participant_normalized": normalize_name(participant_name),
                "row_valid": True,
            }
        )

    if not rows:
        issues.append(
            ValidationIssue(
                "warning",
                "attendance",
                file_name,
                "No attendance rows could be parsed from this CSV file.",
            )
        )
        return empty_df(ATTENDANCE_COLUMNS), issues

    return pd.DataFrame(rows), issues


def aggregate_attendance(raw_attendance_df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate attendance rows to one participant per meeting."""
    if raw_attendance_df.empty:
        return empty_df(ATTENDANCE_AGG_COLUMNS + ["source_file"])

    working = raw_attendance_df.copy()
    working["participant_group_key"] = working.apply(
        lambda row: normalize_email(row["email"]) or normalize_name(row["participant_name"]),
        axis=1,
    )
    grouped = (
        working.groupby(
            ["meeting_key", "meeting_name", "meeting_id", "participant_group_key"],
            dropna=False,
            as_index=False,
        )
        .agg(
            source_file=("source_file", join_unique),
            participant_name=("participant_name", choose_common_value),
            email=("email", choose_common_value),
            total_duration_minutes=("duration_minutes", lambda values: round(pd.Series(values, dtype="float").fillna(0).sum(), 2)),
            first_join=("join_time", safe_min_datetime),
            last_leave=("leave_time", safe_max_datetime),
            join_count=("participant_name", "count"),
        )
    )
    return grouped


# =========================
# Roster and alias loader
# =========================


def prepare_roster_df(
    raw_roster_table: pd.DataFrame,
    name_column: str,
    email_column: str,
) -> tuple[pd.DataFrame, list[ValidationIssue]]:
    """Prepare a normalized roster table."""
    issues: list[ValidationIssue] = []
    if raw_roster_table.empty or not name_column:
        return empty_df(ROSTER_COLUMNS), issues

    prepared = raw_roster_table.copy()
    prepared["roster_name"] = prepared[name_column].map(stringify)
    prepared["roster_email"] = (
        prepared[email_column].map(normalize_email) if email_column and email_column in prepared.columns else ""
    )
    prepared = prepared[
        prepared["roster_name"].astype(str).str.strip().ne("")
        | prepared["roster_email"].astype(str).str.strip().ne("")
    ].copy()
    prepared["roster_name_normalized"] = prepared["roster_name"].map(normalize_name)
    prepared["roster_row_id"] = [f"roster_{idx + 1}" for idx in range(len(prepared))]
    if prepared.empty:
        issues.append(
            ValidationIssue(
                "warning",
                "roster",
                "roster",
                "Roster file loaded but no usable rows remained after cleaning.",
            )
        )
    return prepared[ROSTER_COLUMNS], issues


def prepare_alias_df(raw_alias_table: pd.DataFrame) -> tuple[pd.DataFrame, list[ValidationIssue]]:
    """Prepare a normalized alias mapping table."""
    issues: list[ValidationIssue] = []
    if raw_alias_table.empty:
        return empty_df(ALIAS_COLUMNS), issues

    detected = detect_columns(raw_alias_table.columns, ALIAS_HEADER_SYNONYMS)
    alias_column = detected.get("alias_name")
    canonical_column = detected.get("canonical_name")

    if not alias_column or not canonical_column:
        columns = list(raw_alias_table.columns)
        if len(columns) >= 2:
            alias_column, canonical_column = columns[0], columns[1]
            issues.append(
                ValidationIssue(
                    "warning",
                    "alias",
                    "alias",
                    "Alias headers were not recognized exactly, so the first two columns were used.",
                )
            )
        else:
            issues.append(
                ValidationIssue(
                    "error",
                    "alias",
                    "alias",
                    "Alias file needs at least two columns.",
                )
            )
            return empty_df(ALIAS_COLUMNS), issues

    prepared = raw_alias_table.copy()
    prepared["alias_name"] = prepared[alias_column].map(stringify)
    prepared["canonical_name"] = prepared[canonical_column].map(stringify)
    prepared = prepared[
        prepared["alias_name"].astype(str).str.strip().ne("")
        & prepared["canonical_name"].astype(str).str.strip().ne("")
    ].copy()
    prepared["alias_normalized"] = prepared["alias_name"].map(normalize_name)
    prepared["canonical_normalized"] = prepared["canonical_name"].map(normalize_name)
    return prepared[ALIAS_COLUMNS], issues


# =========================
# Matching helpers
# =========================


def build_candidates(
    roster_df: pd.DataFrame, attendance_agg_df: pd.DataFrame
) -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    """Build global and meeting-level candidate sets for identity matching."""
    if not roster_df.empty:
        global_candidates = roster_df.rename(
            columns={
                "roster_name": "canonical_student_name",
                "roster_email": "candidate_email",
                "roster_name_normalized": "normalized_name",
            }
        ).copy()
        global_candidates["candidate_source"] = "roster"
        return global_candidates[
            ["canonical_student_name", "candidate_email", "normalized_name", "candidate_source"]
        ], {}

    if attendance_agg_df.empty:
        return empty_df(["canonical_student_name", "candidate_email", "normalized_name", "candidate_source"]), {}

    candidates = attendance_agg_df.copy()
    candidates["candidate_email"] = candidates["email"].map(normalize_email)
    candidates["normalized_name"] = candidates["participant_name"].map(normalize_name)
    candidates["canonical_student_name"] = candidates["participant_name"].map(stringify)
    candidates["candidate_source"] = "attendance"
    meeting_map: dict[str, pd.DataFrame] = {}
    for meeting_key, subset in candidates.groupby("meeting_key", dropna=False):
        meeting_map[str(meeting_key)] = subset[
            ["canonical_student_name", "candidate_email", "normalized_name", "candidate_source"]
        ].drop_duplicates()
    return empty_df(["canonical_student_name", "candidate_email", "normalized_name", "candidate_source"]), meeting_map


def build_alias_lookup(alias_df: pd.DataFrame) -> dict[str, list[str]]:
    """Build a normalized alias lookup map."""
    lookup: dict[str, list[str]] = defaultdict(list)
    if alias_df.empty:
        return lookup
    for _, row in alias_df.iterrows():
        alias_norm = normalize_name(row.get("alias_name", ""))
        canonical_norm = normalize_name(row.get("canonical_name", ""))
        if alias_norm and canonical_norm and canonical_norm not in lookup[alias_norm]:
            lookup[alias_norm].append(canonical_norm)
    return lookup


def get_candidates_for_meeting(
    meeting_key: str,
    global_candidates: pd.DataFrame,
    meeting_candidates: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    """Return the candidate universe for a meeting."""
    if not global_candidates.empty:
        return global_candidates.copy()
    return meeting_candidates.get(str(meeting_key), empty_df(["canonical_student_name", "candidate_email", "normalized_name", "candidate_source"])).copy()


def match_record_to_candidates(
    raw_name: str,
    raw_email: str,
    candidates: pd.DataFrame,
    alias_lookup: dict[str, list[str]],
    config: AppConfig,
) -> MatchOutcome:
    """Match a raw display name to canonical candidates conservatively."""
    name_text = stringify(raw_name)
    email_text = normalize_email(raw_email)
    generic_name = looks_generic_device_name(name_text)

    if candidates.empty:
        return MatchOutcome(
            raw_name=name_text,
            canonical_name="",
            match_method="unmatched",
            match_confidence=0.0,
            matched_from_source="none",
            ambiguous=False,
            review_reason="No candidate identities were available for matching.",
        )

    if email_text:
        email_matches = candidates[candidates["candidate_email"].map(normalize_email) == email_text]
        if len(email_matches) == 1:
            candidate = email_matches.iloc[0]
            return MatchOutcome(
                raw_name=name_text,
                canonical_name=stringify(candidate["canonical_student_name"]),
                match_method="exact_email",
                match_confidence=1.0,
                matched_from_source="email",
                ambiguous=False,
                review_reason="",
            )
        if len(email_matches) > 1:
            return MatchOutcome(
                raw_name=name_text,
                canonical_name=stringify(email_matches.iloc[0]["canonical_student_name"]),
                match_method="exact_email",
                match_confidence=1.0,
                matched_from_source="email",
                ambiguous=True,
                review_reason="Multiple candidates share the same email address.",
            )

    exact_matches = candidates[
        candidates["canonical_student_name"].map(stringify).map(str.strip) == name_text.strip()
    ]
    if len(exact_matches) == 1:
        outcome = MatchOutcome(
            raw_name=name_text,
            canonical_name=stringify(exact_matches.iloc[0]["canonical_student_name"]),
            match_method="exact",
            match_confidence=1.0,
            matched_from_source=stringify(exact_matches.iloc[0]["candidate_source"]) or "name",
            ambiguous=False,
            review_reason="",
        )
        if generic_name:
            outcome.review_reason = "Generic device or guest display name requires manual review."
        return outcome
    if len(exact_matches) > 1:
        return MatchOutcome(
            raw_name=name_text,
            canonical_name=stringify(exact_matches.iloc[0]["canonical_student_name"]),
            match_method="exact",
            match_confidence=1.0,
            matched_from_source="name",
            ambiguous=True,
            review_reason="Exact name match is ambiguous because multiple candidates share the same name.",
        )

    normalized_name = normalize_name(name_text)
    if not normalized_name:
        return MatchOutcome(
            raw_name=name_text,
            canonical_name="",
            match_method="unmatched",
            match_confidence=0.0,
            matched_from_source="none",
            ambiguous=False,
            review_reason="Name is blank or generic and could not be matched safely.",
        )

    normalized_matches = candidates[candidates["normalized_name"] == normalized_name]
    if len(normalized_matches) == 1:
        outcome = MatchOutcome(
            raw_name=name_text,
            canonical_name=stringify(normalized_matches.iloc[0]["canonical_student_name"]),
            match_method="normalized_exact",
            match_confidence=0.95,
            matched_from_source=stringify(normalized_matches.iloc[0]["candidate_source"]) or "normalized_name",
            ambiguous=False,
            review_reason="",
        )
        if generic_name:
            outcome.review_reason = "Generic device or guest display name requires manual review."
        return outcome
    if len(normalized_matches) > 1:
        return MatchOutcome(
            raw_name=name_text,
            canonical_name=stringify(normalized_matches.iloc[0]["canonical_student_name"]),
            match_method="normalized_exact",
            match_confidence=0.95,
            matched_from_source="normalized_name",
            ambiguous=True,
            review_reason="Normalized exact match is ambiguous because multiple candidates share the same normalized name.",
        )

    alias_targets = alias_lookup.get(normalized_name, [])
    if alias_targets:
        alias_matches = candidates[candidates["normalized_name"].isin(alias_targets)]
        if len(alias_matches) == 1:
            outcome = MatchOutcome(
                raw_name=name_text,
                canonical_name=stringify(alias_matches.iloc[0]["canonical_student_name"]),
                match_method="alias",
                match_confidence=config.alias_confidence,
                matched_from_source="alias",
                ambiguous=False,
                review_reason="",
            )
            if generic_name:
                outcome.review_reason = "Generic device or guest display name requires manual review."
            return outcome
        if len(alias_matches) > 1:
            return MatchOutcome(
                raw_name=name_text,
                canonical_name=stringify(alias_matches.iloc[0]["canonical_student_name"]),
                match_method="alias",
                match_confidence=config.alias_confidence,
                matched_from_source="alias",
                ambiguous=True,
                review_reason="Alias mapping points to multiple candidate students.",
            )

    choices = {
        idx: normalize_name(row["canonical_student_name"])
        for idx, row in candidates.iterrows()
        if normalize_name(row["canonical_student_name"])
    }
    if not choices:
        return MatchOutcome(
            raw_name=name_text,
            canonical_name="",
            match_method="unmatched",
            match_confidence=0.0,
            matched_from_source="none",
            ambiguous=False,
            review_reason="Candidate list contains no matchable names.",
        )

    fuzzy_results = process.extract(
        normalized_name,
        choices,
        scorer=fuzz.WRatio,
        limit=min(5, len(choices)),
    )
    if not fuzzy_results:
        return MatchOutcome(
            raw_name=name_text,
            canonical_name="",
            match_method="unmatched",
            match_confidence=0.0,
            matched_from_source="none",
            ambiguous=False,
            review_reason="No fuzzy name match was available.",
        )

    _, top_score, top_idx = fuzzy_results[0]
    top_confidence = round(top_score / 100, 4)
    if top_confidence < config.fuzzy_threshold:
        return MatchOutcome(
            raw_name=name_text,
            canonical_name="",
            match_method="unmatched",
            match_confidence=top_confidence,
            matched_from_source="fuzzy",
            ambiguous=False,
            review_reason="Best fuzzy match was below the configured threshold.",
        )

    ambiguous = False
    review_reasons: list[str] = []
    if len(fuzzy_results) > 1:
        _, second_score, _ = fuzzy_results[1]
        if round((top_score - second_score) / 100, 4) < config.ambiguity_margin:
            ambiguous = True
            review_reasons.append("Best fuzzy match was too close to the next candidate.")

    candidate_name = choices[top_idx]
    duplicates = candidates[candidates["normalized_name"] == candidate_name]
    if len(duplicates) > 1:
        ambiguous = True
        review_reasons.append("Multiple candidates share the same normalized name.")

    if generic_name:
        review_reasons.append("Generic device or guest display name requires manual review.")

    top_candidate = candidates.loc[top_idx]
    return MatchOutcome(
        raw_name=name_text,
        canonical_name=stringify(top_candidate["canonical_student_name"]),
        match_method="fuzzy",
        match_confidence=top_confidence,
        matched_from_source=stringify(top_candidate["candidate_source"]) or "fuzzy",
        ambiguous=ambiguous,
        review_reason=" ".join(join_unique(review_reasons, sep=" ").split()),
    )


def evidence_merge_key(
    meeting_key: str,
    evidence_source: str,
    canonical_student_name: str,
    raw_name: str,
    match_method: str,
) -> str:
    """Build a conservative merge key for attendance and transcript evidence."""
    if canonical_student_name and match_method != "unmatched":
        return f"{meeting_key}::canonical::{normalize_name(canonical_student_name)}"
    prefix = "attendance_raw" if evidence_source == "attendance" else "transcript_raw"
    return f"{meeting_key}::{prefix}::{normalize_name(raw_name) or 'unknown'}"


def match_attendance_records(
    attendance_agg_df: pd.DataFrame,
    global_candidates: pd.DataFrame,
    meeting_candidates: dict[str, pd.DataFrame],
    alias_lookup: dict[str, list[str]],
    config: AppConfig,
) -> pd.DataFrame:
    """Match aggregated attendance records to canonical students."""
    if attendance_agg_df.empty:
        return empty_df(MATCHED_COLUMNS)

    records: list[dict[str, Any]] = []
    for _, row in attendance_agg_df.iterrows():
        candidates = get_candidates_for_meeting(row["meeting_key"], global_candidates, meeting_candidates)
        outcome = match_record_to_candidates(
            raw_name=row["participant_name"],
            raw_email=row.get("email", ""),
            candidates=candidates,
            alias_lookup=alias_lookup,
            config=config,
        )
        canonical_name = outcome.canonical_name or stringify(row["participant_name"])
        raw_name = stringify(row["participant_name"])
        records.append(
            {
                "meeting_key": row["meeting_key"],
                "meeting_name": row["meeting_name"],
                "meeting_id": row["meeting_id"],
                "evidence_source": "attendance",
                "source_file": row.get("source_file", ""),
                "canonical_student_name": canonical_name,
                "raw_name": raw_name,
                "raw_attendance_name": raw_name,
                "raw_transcript_speaker": "",
                "email": normalize_email(row.get("email", "")),
                "attendance_minutes": float(row.get("total_duration_minutes", 0) or 0),
                "join_count": int(row.get("join_count", 0) or 0),
                "first_join": row.get("first_join", pd.NaT),
                "last_leave": row.get("last_leave", pd.NaT),
                "speaking_turns": 0,
                "speaking_words": 0,
                "first_speaking_time": "",
                "last_speaking_time": "",
                "speaking_span_minutes": 0.0,
                "match_method": outcome.match_method,
                "match_confidence": float(outcome.match_confidence),
                "matched_from_source": outcome.matched_from_source,
                "ambiguous_match": bool(outcome.ambiguous),
                "identity_review_flag": bool(outcome.review_reason) or outcome.match_confidence < config.safe_auto_approval_threshold,
                "review_reason": outcome.review_reason,
                "raw_name_generic": looks_generic_device_name(raw_name),
                "speaker_is_unknown": False,
                "student_merge_key": evidence_merge_key(
                    row["meeting_key"],
                    "attendance",
                    canonical_name,
                    raw_name,
                    outcome.match_method,
                ),
            }
        )
    return pd.DataFrame(records)


def match_transcript_records(
    speaker_agg_df: pd.DataFrame,
    global_candidates: pd.DataFrame,
    meeting_candidates: dict[str, pd.DataFrame],
    alias_lookup: dict[str, list[str]],
    config: AppConfig,
) -> pd.DataFrame:
    """Match aggregated transcript speakers to canonical students."""
    if speaker_agg_df.empty:
        return empty_df(MATCHED_COLUMNS)

    records: list[dict[str, Any]] = []
    for _, row in speaker_agg_df.iterrows():
        raw_speaker = stringify(row["raw_speaker"]) or UNKNOWN_SPEAKER
        if bool(row.get("speaker_is_unknown", False)):
            outcome = MatchOutcome(
                raw_name=raw_speaker,
                canonical_name="",
                match_method="unmatched",
                match_confidence=0.0,
                matched_from_source="none",
                ambiguous=False,
                review_reason="Transcript contains Unknown Speaker only.",
            )
        else:
            candidates = get_candidates_for_meeting(row["meeting_key"], global_candidates, meeting_candidates)
            outcome = match_record_to_candidates(
                raw_name=raw_speaker,
                raw_email="",
                candidates=candidates,
                alias_lookup=alias_lookup,
                config=config,
            )

        canonical_name = outcome.canonical_name or raw_speaker
        records.append(
            {
                "meeting_key": row["meeting_key"],
                "meeting_name": row["meeting_name"],
                "meeting_id": row["meeting_id"],
                "evidence_source": "transcript",
                "source_file": row.get("source_file", ""),
                "canonical_student_name": canonical_name,
                "raw_name": raw_speaker,
                "raw_attendance_name": "",
                "raw_transcript_speaker": raw_speaker,
                "email": "",
                "attendance_minutes": 0.0,
                "join_count": 0,
                "first_join": pd.NaT,
                "last_leave": pd.NaT,
                "speaking_turns": int(row.get("total_turns", 0) or 0),
                "speaking_words": int(row.get("total_words", 0) or 0),
                "first_speaking_time": row.get("first_speaking_time", ""),
                "last_speaking_time": row.get("last_speaking_time", ""),
                "speaking_span_minutes": float(row.get("speaking_span_minutes", 0) or 0),
                "match_method": outcome.match_method,
                "match_confidence": float(outcome.match_confidence),
                "matched_from_source": outcome.matched_from_source,
                "ambiguous_match": bool(outcome.ambiguous),
                "identity_review_flag": bool(outcome.review_reason) or outcome.match_confidence < config.safe_auto_approval_threshold,
                "review_reason": outcome.review_reason,
                "raw_name_generic": looks_generic_device_name(raw_speaker),
                "speaker_is_unknown": bool(row.get("speaker_is_unknown", False)),
                "student_merge_key": evidence_merge_key(
                    row["meeting_key"],
                    "transcript",
                    canonical_name,
                    raw_speaker,
                    outcome.match_method,
                ),
            }
        )
    return pd.DataFrame(records)


# =========================
# Aggregation and scoring
# =========================


def unique_count_from_joined(text: str) -> int:
    """Count unique items in a pipe-joined string."""
    items = [item.strip() for item in stringify(text).split("|") if item.strip()]
    return len(items)


def generate_row_key(row: pd.Series) -> str:
    """Build a stable review row key."""
    canonical_or_placeholder = stringify(row.get("canonical_student_name", "")) or stringify(
        row.get("raw_attendance_name", "")
    ) or stringify(row.get("raw_transcript_speaker", "")) or "unknown"
    return (
        f"{row.get('meeting_key', '')}|{canonical_or_placeholder}|"
        f"{row.get('raw_attendance_name', '')}|{row.get('raw_transcript_speaker', '')}"
    )


def meeting_unknown_only_map(raw_transcript_df: pd.DataFrame) -> dict[str, bool]:
    """Track meetings whose transcripts contain only unknown speakers."""
    if raw_transcript_df.empty:
        return {}
    status = (
        raw_transcript_df.groupby("meeting_key")["speaker_is_unknown"]
        .agg(lambda values: bool(pd.Series(values).all()))
        .to_dict()
    )
    return {str(key): bool(value) for key, value in status.items()}


def build_final_decision_table(
    matched_df: pd.DataFrame,
    raw_transcript_df: pd.DataFrame,
    roster_present: bool,
    config: AppConfig,
) -> pd.DataFrame:
    """Merge evidence and compute grading recommendations."""
    if matched_df.empty:
        base = empty_df(FINAL_COLUMNS + [
            "meeting_key",
            "weighted_score",
            "final_category",
            "has_attendance_record",
            "has_transcript_record",
            "spoke_strong",
            "matched_from_source",
            "student_merge_key",
            "row_key",
            "unmatched",
            "low_confidence",
            "manual_review_reason",
            "identity_confidence_note",
        ])
        return base

    transcript_unknown_only = meeting_unknown_only_map(raw_transcript_df)

    working = matched_df.copy()
    working["match_confidence"] = pd.to_numeric(working["match_confidence"], errors="coerce").fillna(0.0)

    def choose_first_series(series: pd.Series) -> Any:
        return series.iloc[0] if not series.empty else ""

    grouped = (
        working.groupby(["meeting_key", "student_merge_key"], as_index=False, dropna=False)
        .agg(
            meeting_name=("meeting_name", choose_first_series),
            meeting_id=("meeting_id", choose_first_series),
            canonical_student_name=("canonical_student_name", choose_common_value),
            raw_attendance_name=("raw_attendance_name", join_unique),
            raw_transcript_speaker=("raw_transcript_speaker", join_unique),
            email=("email", choose_common_value),
            attendance_minutes=("attendance_minutes", "sum"),
            speaking_turns=("speaking_turns", "sum"),
            speaking_words=("speaking_words", "sum"),
            speaking_span_minutes=("speaking_span_minutes", "max"),
            join_count=("join_count", "sum"),
            first_join=("first_join", safe_min_datetime),
            last_leave=("last_leave", safe_max_datetime),
            match_method=("match_method", join_unique),
            match_confidence=("match_confidence", "min"),
            matched_from_source=("matched_from_source", join_unique),
            ambiguous_match=("ambiguous_match", "max"),
            identity_review_flag=("identity_review_flag", "max"),
            review_reason=("review_reason", join_unique),
            source_file=("source_file", join_unique),
            has_attendance_record=("evidence_source", lambda values: "attendance" in set(values)),
            has_transcript_record=("evidence_source", lambda values: "transcript" in set(values)),
            speaker_is_unknown=("speaker_is_unknown", "all"),
            raw_name_generic=("raw_name_generic", "max"),
        )
    )

    grouped["attendance_minutes"] = grouped["attendance_minutes"].round(2)
    grouped["speaking_span_minutes"] = grouped["speaking_span_minutes"].round(2)
    grouped["attendance_raw_name_count"] = grouped["raw_attendance_name"].map(unique_count_from_joined)
    grouped["transcript_raw_name_count"] = grouped["raw_transcript_speaker"].map(unique_count_from_joined)
    grouped["attended"] = grouped["attendance_minutes"] >= config.attendance_threshold_minutes
    grouped["spoke"] = (
        (grouped["speaking_words"] >= config.words_threshold)
        | (grouped["speaking_turns"] >= config.turns_threshold)
    )
    if config.span_threshold_enabled:
        grouped["spoke_strong"] = (
            grouped["attended"]
            & grouped["spoke"]
            & (grouped["speaking_span_minutes"] >= config.span_threshold_minutes)
        )
    else:
        grouped["spoke_strong"] = grouped["attended"] & grouped["spoke"]

    grouped["attendance_component"] = (
        (grouped["attendance_minutes"] / max(config.attendance_cap_minutes, 1)).clip(upper=1.0)
        * config.attendance_weight
    )
    grouped["word_component"] = (
        (grouped["speaking_words"] / max(config.word_cap, 1)).clip(upper=1.0)
        * config.word_weight
    )
    grouped["turn_component"] = (
        (grouped["speaking_turns"] / max(config.turn_cap, 1)).clip(upper=1.0)
        * config.turn_weight
    )
    grouped["weighted_score"] = (
        grouped["attendance_component"] + grouped["word_component"] + grouped["turn_component"]
    ).round(4)

    manual_reasons: list[list[str]] = []
    identity_notes: list[str] = []
    for _, row in grouped.iterrows():
        reasons: list[str] = []
        if row["match_confidence"] < config.safe_auto_approval_threshold:
            reasons.append("Low-confidence name match requires manual review.")
        if bool(row["ambiguous_match"]):
            reasons.append("Ambiguous fuzzy or duplicate identity match requires manual review.")
        if bool(row["has_transcript_record"]) and not bool(row["has_attendance_record"]):
            reasons.append("Speaking evidence present but attendance match missing.")
        if bool(row["has_attendance_record"]) and transcript_unknown_only.get(str(row["meeting_key"]), False):
            reasons.append("Attendance exists but transcript only shows Unknown Speaker.")
        if not stringify(row["canonical_student_name"]):
            reasons.append("Missing canonical student identity requires manual review.")
        if bool(row["raw_name_generic"]):
            reasons.append("Generic device or guest display name requires manual review.")
        if int(row["attendance_raw_name_count"]) > 1 or int(row["transcript_raw_name_count"]) > 1:
            reasons.append("Duplicate canonical match conflict in the same meeting requires manual review.")
        existing_reason = stringify(row["review_reason"])
        if existing_reason:
            reasons.append(existing_reason)
        manual_reasons.append(list(dict.fromkeys([reason for reason in reasons if reason])))
        if not roster_present and bool(row["has_attendance_record"]):
            identity_notes.append("Roster not uploaded; identity confidence relies on uploaded Zoom data.")
        else:
            identity_notes.append("")

    grouped["manual_review_reason"] = [" | ".join(reasons) for reasons in manual_reasons]
    grouped["identity_confidence_note"] = identity_notes
    grouped["manual_review"] = grouped["manual_review_reason"].astype(str).str.strip().ne("")
    grouped["unmatched"] = grouped["match_method"].astype(str).str.contains("unmatched", case=False, na=False)
    grouped["low_confidence"] = grouped["match_confidence"] < config.safe_auto_approval_threshold
    grouped["reviewer_notes"] = ""
    grouped["reviewer_override"] = ""
    grouped["recommend_award"] = False

    grouped = recompute_decisions(grouped, config, overrides=None)
    grouped["row_key"] = grouped.apply(generate_row_key, axis=1)

    final_columns = FINAL_COLUMNS + [
        "meeting_key",
        "weighted_score",
        "final_category",
        "has_attendance_record",
        "has_transcript_record",
        "spoke_strong",
        "matched_from_source",
        "student_merge_key",
        "row_key",
        "unmatched",
        "low_confidence",
        "manual_review_reason",
        "identity_confidence_note",
        "source_file",
    ]
    for column in final_columns:
        if column not in grouped.columns:
            grouped[column] = ""
    return grouped[final_columns]


def policy_recommendation(row: pd.Series, config: AppConfig) -> tuple[bool, str]:
    """Apply the selected bonus policy to a row."""
    mode = config.bonus_policy_mode
    if mode == "attended only":
        if bool(row["attended"]):
            return True, "Attendance-only policy selected and attendance threshold met."
        return False, "Attendance-only policy selected but attendance threshold was not met."

    if mode == "spoke only":
        if bool(row["has_attendance_record"]) and bool(row["spoke"]):
            return True, "Spoke-only policy selected and speaking threshold met with attendance evidence present."
        if not bool(row["has_attendance_record"]):
            return False, "Speaking evidence present but attendance match missing."
        return False, "Spoke-only policy selected but speaking threshold was not met."

    if mode == "attended and spoke":
        if bool(row["attended"]) and bool(row["spoke"]):
            return True, "Met attendance threshold and speaking threshold."
        if bool(row["attended"]) and not bool(row["spoke"]):
            return False, "Attendance met but speaking below threshold."
        if bool(row["spoke"]) and not bool(row["attended"]):
            return False, "Speaking evidence present but attendance threshold was not met."
        return False, "Attendance and speaking thresholds were not both met."

    weighted_score = float(row.get("weighted_score", 0) or 0)
    if bool(row["has_attendance_record"]) and weighted_score >= config.weighted_threshold:
        return True, f"Weighted score policy selected and score {weighted_score:.2f} met the configured threshold."
    if not bool(row["has_attendance_record"]):
        return False, "Weighted score policy requires attendance evidence, but attendance match is missing."
    return False, f"Weighted score {weighted_score:.2f} was below the configured threshold."


def recompute_decisions(
    final_df: pd.DataFrame,
    config: AppConfig,
    overrides: Optional[dict[str, dict[str, Any]]],
) -> pd.DataFrame:
    """Recompute final recommendation fields after overrides."""
    updated = final_df.copy()
    final_categories: list[str] = []
    recommend_awards: list[bool] = []
    reasons: list[str] = []

    for _, row in updated.iterrows():
        row_key = stringify(row.get("row_key", ""))
        override_state = overrides.get(row_key, {}) if overrides else {}
        reviewer_override = stringify(row.get("reviewer_override", ""))
        reviewer_notes = stringify(row.get("reviewer_notes", ""))

        policy_award, policy_reason = policy_recommendation(row, config)
        category = "Award" if policy_award else "Do_Not_Award"
        recommend_award = bool(row.get("recommend_award", policy_award))
        decision_reason = policy_reason

        if "recommend_award" in override_state and not bool(row.get("manual_review", False)):
            category = "Award" if bool(row["recommend_award"]) else "Do_Not_Award"
            recommend_award = bool(row["recommend_award"])
            decision_reason = "Reviewer updated the award recommendation."

        if bool(row.get("manual_review", False)):
            category = "Manual_Review"
            recommend_award = False
            decision_reason = stringify(row.get("manual_review_reason", "")) or "Manual review required."

        if reviewer_override in {"Award", "Do_Not_Award", "Manual_Review"}:
            category = reviewer_override
            recommend_award = reviewer_override == "Award"
            decision_reason = f"Reviewer override applied: {reviewer_override.replace('_', ' ')}."

        if reviewer_notes:
            decision_reason = f"{decision_reason} Notes: {reviewer_notes}".strip()
        identity_note = stringify(row.get("identity_confidence_note", ""))
        if identity_note and identity_note not in decision_reason:
            decision_reason = f"{decision_reason} {identity_note}".strip()

        final_categories.append(category)
        recommend_awards.append(recommend_award)
        reasons.append(decision_reason)

    updated["final_category"] = final_categories
    updated["recommend_award"] = recommend_awards
    updated["decision_reason"] = reasons
    return updated


# =========================
# Session-state override helpers
# =========================


EDITABLE_FIELDS = [
    "canonical_student_name",
    "attended",
    "spoke",
    "recommend_award",
    "manual_review",
    "reviewer_notes",
    "reviewer_override",
]


def get_override_store() -> dict[str, dict[str, Any]]:
    """Return the session-state override store."""
    if "manual_overrides" not in st.session_state:
        st.session_state["manual_overrides"] = {}
    return st.session_state["manual_overrides"]


def apply_manual_overrides(
    final_df: pd.DataFrame, config: AppConfig
) -> pd.DataFrame:
    """Apply persisted overrides and recompute final decisions."""
    if final_df.empty:
        return final_df.copy()
    updated = final_df.copy()
    override_store = get_override_store()
    for idx, row in updated.iterrows():
        row_key = stringify(row.get("row_key", ""))
        row_override = override_store.get(row_key, {})
        for field, value in row_override.items():
            if field in updated.columns:
                updated.at[idx, field] = value
    return recompute_decisions(updated, config, override_store)


def persist_editor_overrides(
    base_editor_df: pd.DataFrame, edited_editor_df: pd.DataFrame
) -> bool:
    """Persist changed review fields from the data editor into session state."""
    overrides = get_override_store()
    changed = False
    for row_key in edited_editor_df.index:
        current = edited_editor_df.loc[row_key]
        original = base_editor_df.loc[row_key]
        row_override = overrides.get(row_key, {}).copy()
        for field in EDITABLE_FIELDS:
            new_value = current[field]
            old_value = original[field]
            if pd.isna(new_value):
                new_value = ""
            if pd.isna(old_value):
                old_value = ""
            if isinstance(new_value, str):
                new_value = new_value.strip()
            if isinstance(old_value, str):
                old_value = old_value.strip()
            if new_value != old_value:
                row_override[field] = new_value
            elif field in row_override:
                row_override.pop(field, None)

        if row_override:
            if overrides.get(row_key) != row_override:
                overrides[row_key] = row_override
                changed = True
        elif row_key in overrides:
            overrides.pop(row_key, None)
            changed = True

    st.session_state["manual_overrides"] = overrides
    return changed


# =========================
# Filtering helpers
# =========================


def apply_filters(df: pd.DataFrame, config: AppConfig) -> pd.DataFrame:
    """Apply sidebar filters to a table."""
    if df.empty:
        return df.copy()

    filtered = df.copy()
    if not config.combine_all_meetings and config.selected_meetings:
        filtered = filtered[filtered["meeting_name"].isin(config.selected_meetings)]

    if config.search_text:
        search = config.search_text.lower()
        mask = (
            filtered["canonical_student_name"].astype(str).str.lower().str.contains(search, na=False)
            | filtered["raw_attendance_name"].astype(str).str.lower().str.contains(search, na=False)
            | filtered["raw_transcript_speaker"].astype(str).str.lower().str.contains(search, na=False)
        )
        filtered = filtered[mask]

    category_filters = []
    if config.show_only_award:
        category_filters.append("Award")
    if config.show_only_do_not_award:
        category_filters.append("Do_Not_Award")
    if config.show_only_manual_review:
        category_filters.append("Manual_Review")
    if category_filters and "final_category" in filtered.columns:
        filtered = filtered[filtered["final_category"].isin(category_filters)]

    if config.show_only_unmatched and "unmatched" in filtered.columns:
        filtered = filtered[filtered["unmatched"] == True]  # noqa: E712

    if config.show_only_low_confidence and "low_confidence" in filtered.columns:
        filtered = filtered[filtered["low_confidence"] == True]  # noqa: E712

    return filtered


# =========================
# Excel export writer
# =========================


def autosize_worksheet(worksheet: Any) -> None:
    """Adjust worksheet column widths with a practical upper bound."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def generate_audit_df(
    issues: list[ValidationIssue],
    uploaded_filenames: Sequence[str],
    config: AppConfig,
    final_df: pd.DataFrame,
) -> pd.DataFrame:
    """Create the audit log sheet."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows: list[dict[str, str]] = []
    rows.append(
        {
            "timestamp": now,
            "category": "export",
            "detail": f"Workbook generated for files: {', '.join(uploaded_filenames) if uploaded_filenames else 'None'}",
        }
    )
    rows.append(
        {
            "timestamp": now,
            "category": "policy",
            "detail": (
                f"Policy={config.bonus_policy_mode}; attendance_threshold={config.attendance_threshold_minutes}; "
                f"words_threshold={config.words_threshold}; turns_threshold={config.turns_threshold}; "
                f"span_enabled={config.span_threshold_enabled}; span_threshold={config.span_threshold_minutes}; "
                f"fuzzy_threshold={config.fuzzy_threshold}; safe_auto_threshold={config.safe_auto_approval_threshold}"
            ),
        }
    )

    if not final_df.empty:
        counts = final_df["final_category"].value_counts().to_dict()
        rows.append(
            {
                "timestamp": now,
                "category": "summary",
                "detail": (
                    f"Award={counts.get('Award', 0)}, "
                    f"Do_Not_Award={counts.get('Do_Not_Award', 0)}, "
                    f"Manual_Review={counts.get('Manual_Review', 0)}"
                ),
            }
        )
        unmatched_rows = final_df[final_df["unmatched"] == True]  # noqa: E712
        for _, row in unmatched_rows.iterrows():
            rows.append(
                {
                    "timestamp": now,
                    "category": "unmatched",
                    "detail": (
                        f"{row['meeting_name']}: unmatched identity for "
                        f"{row['raw_attendance_name'] or row['raw_transcript_speaker']}"
                    ),
                }
            )
        low_confidence_rows = final_df[final_df["low_confidence"] == True]  # noqa: E712
        for _, row in low_confidence_rows.iterrows():
            rows.append(
                {
                    "timestamp": now,
                    "category": "low_confidence",
                    "detail": (
                        f"{row['meeting_name']}: {row['canonical_student_name']} matched with confidence "
                        f"{row['match_confidence']:.2f}"
                    ),
                }
            )

    for issue in issues:
        rows.append(
            {
                "timestamp": now,
                "category": issue.level,
                "detail": f"[{issue.source_type}] {issue.source_file}: {issue.message} {issue.row_hint}".strip(),
            }
        )

    return pd.DataFrame(rows)


def build_excel_workbook(
    config: AppConfig,
    raw_transcript_df: pd.DataFrame,
    raw_attendance_df: pd.DataFrame,
    speaker_agg_df: pd.DataFrame,
    attendance_agg_df: pd.DataFrame,
    matched_df: pd.DataFrame,
    final_df: pd.DataFrame,
    issues: list[ValidationIssue],
    uploaded_filenames: Sequence[str],
) -> bytes:
    """Create an in-memory Excel workbook for download."""
    config_rows = [
        {"Setting": "Export Timestamp", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Setting": "Bonus Policy Mode", "Value": config.bonus_policy_mode},
        {"Setting": "Attendance Threshold Minutes", "Value": config.attendance_threshold_minutes},
        {"Setting": "Word Threshold", "Value": config.words_threshold},
        {"Setting": "Turn Threshold", "Value": config.turns_threshold},
        {"Setting": "Span Threshold Enabled", "Value": config.span_threshold_enabled},
        {"Setting": "Span Threshold Minutes", "Value": config.span_threshold_minutes},
        {"Setting": "Fuzzy Threshold", "Value": config.fuzzy_threshold},
        {"Setting": "Safe Auto-Approval Threshold", "Value": config.safe_auto_approval_threshold},
        {"Setting": "Alias Confidence", "Value": config.alias_confidence},
        {"Setting": "Attendance Weight", "Value": config.attendance_weight},
        {"Setting": "Word Weight", "Value": config.word_weight},
        {"Setting": "Turn Weight", "Value": config.turn_weight},
        {"Setting": "Attendance Cap Minutes", "Value": config.attendance_cap_minutes},
        {"Setting": "Word Cap", "Value": config.word_cap},
        {"Setting": "Turn Cap", "Value": config.turn_cap},
        {"Setting": "Weighted Threshold", "Value": config.weighted_threshold},
        {"Setting": "Uploaded Files", "Value": ", ".join(uploaded_filenames)},
    ]
    config_df = pd.DataFrame(config_rows)
    audit_df = generate_audit_df(issues, uploaded_filenames, config, final_df)

    award_df = final_df[final_df["final_category"] == "Award"].copy()
    do_not_award_df = final_df[final_df["final_category"] == "Do_Not_Award"].copy()
    manual_review_df = final_df[final_df["final_category"] == "Manual_Review"].copy()

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_tables = {
            CONFIG_SHEET_NAME: config_df,
            RAW_TRANSCRIPT_SHEET: raw_transcript_df.drop(columns=["start_seconds", "end_seconds", "interval_key"], errors="ignore"),
            RAW_ATTENDANCE_SHEET: raw_attendance_df,
            AGG_SPEAKERS_SHEET: speaker_agg_df.drop(columns=["first_speaking_seconds", "last_speaking_seconds", "speaker_is_unknown"], errors="ignore"),
            AGG_ATTENDANCE_SHEET: attendance_agg_df,
            MATCHED_SHEET: matched_df,
            AWARD_SHEET: award_df,
            DO_NOT_AWARD_SHEET: do_not_award_df,
            MANUAL_REVIEW_SHEET: manual_review_df,
            AUDIT_SHEET: audit_df,
        }
        for sheet_name, table in export_tables.items():
            safe_table = table.copy()
            safe_table.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.book[sheet_name]
            worksheet.freeze_panes = "A2"
            if worksheet.max_row > 1 and worksheet.max_column > 0:
                worksheet.auto_filter.ref = worksheet.dimensions
            autosize_worksheet(worksheet)
    output.seek(0)
    return output.getvalue()


# =========================
# Sample data generator
# =========================


def generate_sample_files() -> dict[str, bytes]:
    """Generate in-memory sample files for instructors to test the app."""
    transcript = """WEBVTT

00:00:01.000 --> 00:00:05.000
<v Jane Doe>Hi everyone, thanks for joining the review session.

00:00:07.000 --> 00:00:10.000
<v John Smith>I have a question about the assignment rubric.

00:00:12.000 --> 00:00:15.000
<v Jane Doe>Great question, let's look at the example together.

00:00:17.000 --> 00:00:19.000
Unknown Speaker: Sounds good.
"""

    attendance = """Topic,Meeting ID,Name (Original Name),User Email,Join Time,Leave Time,Duration (Minutes)
BIO101 Week 4,987654321,Jane Doe,jane@example.edu,2026-03-20 09:00:00,2026-03-20 09:58:00,58
BIO101 Week 4,987654321,John Smith,john@example.edu,2026-03-20 09:03:00,2026-03-20 09:55:00,52
BIO101 Week 4,987654321,Guest,,2026-03-20 09:10:00,2026-03-20 09:20:00,10
"""

    roster = """Student Name,Email Address
Jane Doe,jane@example.edu
John Smith,john@example.edu
Maria Lopez,maria@example.edu
"""

    alias = """alias_name,canonical_name
J. Doe,Jane Doe
Johnny,John Smith
"""

    return {
        "BIO101_Week_4_987654321_transcript.vtt": transcript.encode("utf-8"),
        "sample_participants.csv": attendance.encode("utf-8"),
        "sample_roster.csv": roster.encode("utf-8"),
        "sample_aliases.csv": alias.encode("utf-8"),
    }


# =========================
# Streamlit UI helpers
# =========================


def render_global_styles() -> None:
    """Inject a polished visual system for the app."""
    st.markdown(
        """
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&display=swap');

            :root {
                --zoom-bg: #080d1a;
                --zoom-bg-deep: #0b1225;
                --zoom-panel: rgba(17, 25, 50, 0.94);
                --zoom-panel-2: rgba(23, 33, 66, 0.96);
                --zoom-panel-3: rgba(33, 47, 89, 0.92);
                --zoom-line: rgba(118, 142, 209, 0.16);
                --zoom-line-strong: rgba(118, 142, 209, 0.28);
                --zoom-text: #f4f7ff;
                --zoom-text-strong: #ffffff;
                --zoom-muted: #a4b0d6;
                --zoom-subtle: #7c89b0;
                --zoom-primary: #2d8cff;
                --zoom-primary-strong: #0b5cff;
                --zoom-cyan: #4cc9ff;
                --zoom-success: #4da6ff;
                --zoom-success-soft: rgba(77, 166, 255, 0.14);
                --zoom-warning: #f59e0b;
                --zoom-warning-soft: rgba(245, 158, 11, 0.14);
                --zoom-danger: #f45d6c;
                --zoom-danger-soft: rgba(244, 93, 108, 0.14);
                --zoom-shadow: 0 24px 60px rgba(1, 6, 20, 0.46);
            }

            html, body, [class*="css"] {
                color-scheme: dark;
            }

            .stApp {
                --primary-color: var(--zoom-primary);
                --background-color: var(--zoom-bg);
                --secondary-background-color: var(--zoom-panel);
                --text-color: var(--zoom-text);
                background:
                    radial-gradient(circle at top left, rgba(45, 140, 255, 0.16), transparent 26%),
                    radial-gradient(circle at top right, rgba(76, 201, 255, 0.12), transparent 22%),
                    linear-gradient(180deg, #060b16 0%, #0a1122 44%, #0b1225 100%);
                color: var(--zoom-text);
            }

            .main .block-container {
                max-width: 1280px;
                padding-top: 2.25rem;
                padding-bottom: 5.75rem;
            }

            .stApp, .stMarkdown, .stText, p, li, div, span, label {
                font-family: "Manrope", "Avenir Next", "Trebuchet MS", sans-serif;
            }

            input,
            textarea,
            select {
                accent-color: var(--zoom-primary) !important;
            }

            h1, h2, h3, h4, .hero-title, .section-title, .card-title {
                font-family: "Manrope", "Avenir Next", "Trebuchet MS", sans-serif;
                letter-spacing: -0.03em;
            }

            section[data-testid="stSidebar"] {
                --primary-color: var(--zoom-primary);
                background:
                    linear-gradient(180deg, rgba(7, 13, 29, 0.98), rgba(11, 18, 37, 0.99));
                border-right: 1px solid rgba(118, 142, 209, 0.14);
            }

            section[data-testid="stSidebar"] * {
                color: var(--zoom-text);
            }

            section[data-testid="stSidebar"] .stSlider label,
            section[data-testid="stSidebar"] .stTextInput label,
            section[data-testid="stSidebar"] .stSelectbox label,
            section[data-testid="stSidebar"] .stMultiSelect label,
            section[data-testid="stSidebar"] .stToggle label {
                color: var(--zoom-text);
                font-weight: 600;
            }

            section[data-testid="stSidebar"] .stExpander {
                border: 1px solid rgba(118, 142, 209, 0.14);
                border-radius: 20px;
                background: rgba(255, 255, 255, 0.02);
                overflow: hidden;
            }

            .sidebar-panel {
                background: rgba(15, 22, 45, 0.82) !important;
                border-color: rgba(118, 142, 209, 0.14) !important;
                box-shadow: none !important;
                padding: 1rem 1rem 0.95rem 1rem !important;
            }

            .sidebar-panel .card-kicker {
                color: rgba(220, 233, 255, 0.72) !important;
            }

            .sidebar-panel h3 {
                margin: 0.2rem 0 0.45rem 0 !important;
                color: var(--zoom-text-strong) !important;
            }

            .sidebar-panel p {
                margin: 0 !important;
                color: rgba(220, 233, 255, 0.82) !important;
                line-height: 1.6 !important;
            }

            .saas-shell {
                background: var(--zoom-panel);
                border: 1px solid var(--zoom-line);
                border-radius: 26px;
                box-shadow: var(--zoom-shadow);
                padding: 1.2rem 1.2rem 1.05rem 1.2rem;
            }

            .hero-shell {
                background:
                    radial-gradient(circle at top right, rgba(76, 201, 255, 0.18), transparent 28%),
                    linear-gradient(180deg, rgba(16, 24, 48, 0.98), rgba(17, 28, 56, 0.98));
                border: 1px solid var(--zoom-line-strong);
                border-radius: 32px;
                color: var(--zoom-text-strong);
                padding: 2rem 2rem 1.8rem 2rem;
                box-shadow: var(--zoom-shadow);
                min-height: 100%;
            }

            .hero-kicker,
            .section-kicker,
            .card-kicker,
            .metric-label,
            .pill {
                text-transform: uppercase;
                letter-spacing: 0.14em;
                font-size: 0.72rem;
                font-weight: 800;
            }

            .hero-kicker {
                color: rgba(244, 247, 255, 0.68);
                margin-bottom: 0.85rem;
            }

            .hero-title {
                font-size: clamp(2.3rem, 4vw, 4rem);
                line-height: 1.02;
                margin: 0 0 0.85rem 0;
            }

            .hero-copy {
                color: var(--zoom-muted);
                font-size: 1.02rem;
                line-height: 1.72;
                margin-bottom: 1.15rem;
                max-width: 46rem;
            }

            .pill-row {
                display: flex;
                flex-wrap: wrap;
                gap: 0.65rem;
                margin-top: 1rem;
            }

            .pill {
                display: inline-flex;
                align-items: center;
                gap: 0.35rem;
                padding: 0.58rem 0.9rem;
                border-radius: 999px;
                background: rgba(45, 140, 255, 0.10);
                color: #dce9ff;
                border: 1px solid rgba(76, 201, 255, 0.16);
            }

            .hero-panel,
            .info-card,
            .workflow-card,
            .metric-card,
            .notice-card,
            .guide-card,
            .sample-card,
            .workspace-banner,
            .status-strip {
                background: var(--zoom-panel);
                border: 1px solid var(--zoom-line);
                border-radius: 24px;
                box-shadow: var(--zoom-shadow);
            }

            .hero-panel {
                padding: 1.25rem 1.25rem 1.1rem 1.25rem;
                min-height: 100%;
            }

            .hero-panel h3,
            .card-title {
                margin: 0 0 0.55rem 0;
                color: var(--zoom-text-strong);
                font-size: 1.14rem;
                font-weight: 700;
            }

            .hero-panel p,
            .hero-panel li,
            .card-copy,
            .section-copy {
                color: var(--zoom-muted);
                line-height: 1.68;
                margin: 0;
            }

            .hero-list,
            .sheet-list,
            .guide-list {
                list-style: none;
                padding: 0;
                margin: 0.95rem 0 0 0;
            }

            .hero-list li,
            .sheet-list li,
            .guide-list li {
                display: flex;
                gap: 0.9rem;
                align-items: flex-start;
                justify-content: space-between;
                padding: 0.78rem 0;
                border-bottom: 1px solid rgba(118, 142, 209, 0.10);
                color: var(--zoom-text);
                font-size: 0.95rem;
            }

            .hero-list li:last-child,
            .sheet-list li:last-child,
            .guide-list li:last-child {
                border-bottom: none;
            }

            .hero-list span:last-child,
            .sheet-list span:last-child {
                color: var(--zoom-muted);
                text-align: right;
            }

            .section-shell {
                margin: 0.25rem 0 1.85rem 0;
            }

            .section-kicker {
                color: var(--zoom-cyan);
                margin-bottom: 0.45rem;
            }

            .section-title {
                color: var(--zoom-text-strong);
                font-size: clamp(1.5rem, 2.5vw, 2.2rem);
                margin: 0 0 0.4rem 0;
                font-weight: 800;
            }

            .section-copy {
                max-width: 56rem;
                color: var(--zoom-muted);
            }

            .info-card,
            .workflow-card,
            .metric-card,
            .notice-card,
            .sample-card {
                padding: 1.18rem 1.18rem 1.08rem 1.18rem;
                min-height: 100%;
            }

            .guide-card {
                padding: 1.2rem 1.25rem 1.15rem 1.25rem;
            }

            .info-card,
            .workflow-card,
            .metric-card,
            .notice-card,
            .guide-card,
            .sample-card {
                display: flex;
                flex-direction: column;
                gap: 0.38rem;
                height: 100%;
            }

            .guide-step,
            .workflow-step {
                width: 1.95rem;
                height: 1.95rem;
                min-width: 1.95rem;
                border-radius: 999px;
                background: rgba(45, 140, 255, 0.14);
                color: #dce9ff;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                font-weight: 800;
                font-size: 0.88rem;
            }

            .workflow-step {
                margin-bottom: 0.95rem;
            }

            .status-strip {
                padding: 1.15rem 1.15rem 1.05rem 1.15rem;
            }

            .status-grid {
                display: grid;
                grid-template-columns: repeat(4, minmax(0, 1fr));
                gap: 0.9rem;
            }

            .status-chip {
                border-radius: 18px;
                padding: 0.95rem 1rem;
                border: 1px solid rgba(118, 142, 209, 0.12);
                background: var(--zoom-panel-2);
            }

            .status-chip strong {
                display: block;
                color: var(--zoom-text-strong);
                margin-bottom: 0.25rem;
                font-size: 0.97rem;
            }

            .status-chip span {
                color: var(--zoom-muted);
                font-size: 0.9rem;
                line-height: 1.52;
            }

            .status-chip.ok {
                background: linear-gradient(180deg, rgba(77, 166, 255, 0.16), rgba(17, 29, 45, 0.96));
            }

            .status-chip.pending {
                background: linear-gradient(180deg, rgba(45, 140, 255, 0.10), rgba(17, 25, 50, 0.96));
            }

            .metric-card {
                padding: 1.05rem 1.05rem 1rem 1.05rem;
            }

            .metric-card.metric-award {
                background: linear-gradient(180deg, rgba(77, 166, 255, 0.20), rgba(17, 25, 50, 0.96));
            }

            .metric-card.metric-warning {
                background: linear-gradient(180deg, rgba(245, 158, 11, 0.18), rgba(17, 25, 50, 0.96));
            }

            .metric-card.metric-review {
                background: linear-gradient(180deg, rgba(244, 93, 108, 0.18), rgba(17, 25, 50, 0.96));
            }

            .metric-card.metric-neutral {
                background: linear-gradient(180deg, rgba(45, 140, 255, 0.12), rgba(17, 25, 50, 0.96));
            }

            .metric-label {
                color: var(--zoom-subtle);
                margin-bottom: 0.58rem;
            }

            .metric-value {
                color: var(--zoom-text-strong);
                font-size: 2rem;
                font-weight: 800;
                line-height: 1;
            }

            .metric-footnote {
                color: var(--zoom-muted);
                font-size: 0.88rem;
                margin-top: 0.58rem;
                line-height: 1.5;
            }

            .notice-card.notice-success {
                background: linear-gradient(180deg, rgba(77, 166, 255, 0.17), rgba(17, 25, 50, 0.96));
                border-color: rgba(77, 166, 255, 0.24);
            }

            .notice-card.notice-warning {
                background: linear-gradient(180deg, rgba(245, 158, 11, 0.16), rgba(17, 25, 50, 0.96));
                border-color: rgba(245, 158, 11, 0.20);
            }

            .notice-card.notice-danger {
                background: linear-gradient(180deg, rgba(244, 93, 108, 0.16), rgba(17, 25, 50, 0.96));
                border-color: rgba(244, 93, 108, 0.20);
            }

            .notice-title {
                font-size: 1rem;
                font-weight: 700;
                color: var(--zoom-text-strong);
                margin-bottom: 0.32rem;
            }

            .notice-copy {
                color: var(--zoom-muted);
                margin: 0;
                line-height: 1.56;
            }

            .workspace-banner {
                background: linear-gradient(180deg, rgba(33, 47, 89, 0.76), rgba(17, 25, 50, 0.96));
                padding: 1rem 1.05rem;
            }

            .workspace-banner strong {
                color: var(--zoom-text-strong);
            }

            .sample-card {
                padding: 0;
                overflow: hidden;
                gap: 0;
            }

            .sample-card-body {
                padding: 1.2rem 1.2rem 1rem 1.2rem;
                flex: 1;
                background: linear-gradient(180deg, rgba(17, 25, 50, 0.98), rgba(14, 22, 44, 0.98));
            }

            .sample-card-file {
                margin: 0.25rem 0 0 0;
                color: var(--zoom-text-strong);
                font-size: 1rem;
                line-height: 1.32;
                font-weight: 700;
                word-break: break-word;
            }

            .sample-card-copy {
                margin: 1rem 0 0 0;
                color: var(--zoom-muted);
                line-height: 1.68;
            }

            .sample-card-cta {
                display: flex;
                align-items: center;
                justify-content: center;
                min-height: 3.55rem;
                padding: 0.9rem 1rem;
                border-top: 1px solid rgba(186, 227, 255, 0.14);
                background: linear-gradient(180deg, #2d8cff 0%, #0b5cff 100%);
                color: #ffffff !important;
                font-weight: 800;
                text-decoration: none;
                text-align: center;
            }

            .sample-card-cta:hover {
                background: linear-gradient(180deg, #4da6ff 0%, #1869ff 100%);
            }

            .stSelectbox [data-baseweb="select"] > div,
            .stMultiSelect [data-baseweb="select"] > div,
            .stTextInput input,
            .stNumberInput input,
            .stTextArea textarea {
                background: var(--zoom-panel-2) !important;
                color: var(--zoom-text) !important;
                border-radius: 14px !important;
                border: 1px solid var(--zoom-line) !important;
            }

            .stSelectbox [data-baseweb="select"]:focus-within > div,
            .stMultiSelect [data-baseweb="select"]:focus-within > div,
            .stTextInput input:focus,
            .stNumberInput input:focus,
            .stTextArea textarea:focus {
                border-color: rgba(76, 201, 255, 0.34) !important;
                box-shadow: 0 0 0 4px rgba(45, 140, 255, 0.14) !important;
            }

            div[data-baseweb="popover"] > div,
            ul[role="listbox"] {
                background: var(--zoom-panel-2) !important;
                color: var(--zoom-text) !important;
                border: 1px solid var(--zoom-line) !important;
            }

            div[data-testid="stFileUploader"] {
                background: var(--zoom-panel);
                border: 1px solid var(--zoom-line);
                border-radius: 24px;
                padding: 0.55rem 0.9rem 0.95rem 0.9rem;
            }

            div[data-testid="stFileUploader"] section {
                border: 2px dashed rgba(76, 201, 255, 0.22);
                border-radius: 18px;
                background: rgba(45, 140, 255, 0.07);
            }

            .stSlider {
                padding-top: 0.15rem;
            }

            .stSlider [data-baseweb="slider"] {
                padding-top: 0.3rem;
                padding-bottom: 0.3rem;
            }

            .stSlider [data-baseweb="slider"] > div {
                background: transparent !important;
            }

            .stSlider [data-baseweb="slider"] > div > div {
                background: rgba(118, 142, 209, 0.28) !important;
            }

            .stSlider [data-baseweb="slider"] > div > div > div {
                background: linear-gradient(90deg, rgba(76, 201, 255, 0.94), rgba(45, 140, 255, 0.98)) !important;
            }

            .stSlider [role="slider"] {
                background: radial-gradient(circle at 35% 35%, #ffffff 0%, #dce9ff 30%, #75c7ff 58%, #2d8cff 100%) !important;
                border: 2px solid rgba(205, 234, 255, 0.88) !important;
                box-shadow: 0 0 0 5px rgba(45, 140, 255, 0.22) !important;
            }

            .stSlider [role="slider"]:focus,
            .stSlider [role="slider"]:focus-visible {
                outline: none !important;
                box-shadow: 0 0 0 6px rgba(76, 201, 255, 0.26) !important;
            }

            .stSlider div[data-testid="stTickBarMin"],
            .stSlider div[data-testid="stTickBarMax"],
            .stSlider div[data-testid*="TickBar"] {
                background: rgba(77, 166, 255, 0.82) !important;
            }

            .stSlider [data-testid="stThumbValue"],
            .stSlider p {
                color: var(--zoom-muted) !important;
            }

            .stToggle [data-baseweb="switch"] > div {
                background: rgba(118, 142, 209, 0.24) !important;
            }

            .stToggle [data-baseweb="switch"] input:checked + div,
            .stToggle [data-baseweb="switch"] div[aria-checked="true"] {
                background: linear-gradient(90deg, #2d8cff, #4cc9ff) !important;
            }

            .stDownloadButton > button,
            .stButton > button,
            div[data-testid="stFileUploader"] button,
            button[data-testid="baseButton-secondary"],
            button[data-testid="baseButton-primary"] {
                border-radius: 16px !important;
                border: 1px solid rgba(76, 201, 255, 0.18) !important;
                background: linear-gradient(180deg, #2d8cff 0%, #0b5cff 100%) !important;
                color: #ffffff !important;
                font-weight: 800 !important;
                min-height: 2.85rem !important;
                padding: 0.78rem 1rem !important;
                box-shadow: 0 14px 34px rgba(11, 92, 255, 0.34) !important;
                transition: transform 120ms ease, box-shadow 120ms ease, border-color 120ms ease;
            }

            .stDownloadButton > button:hover,
            .stButton > button:hover,
            div[data-testid="stFileUploader"] button:hover,
            button[data-testid="baseButton-secondary"]:hover,
            button[data-testid="baseButton-primary"]:hover {
                border-color: rgba(76, 201, 255, 0.32) !important;
                box-shadow: 0 18px 36px rgba(11, 92, 255, 0.40) !important;
                transform: translateY(-1px);
                color: #ffffff !important;
            }

            div[data-baseweb="tab-list"] {
                gap: 0.45rem;
                background: rgba(17, 25, 50, 0.88);
                border: 1px solid var(--zoom-line);
                padding: 0.38rem;
                border-radius: 18px;
            }

            div[data-baseweb="tab-list"] button {
                border-radius: 14px;
                color: var(--zoom-muted) !important;
                font-weight: 700;
                background: transparent;
            }

            div[data-baseweb="tab-list"] button[aria-selected="true"] {
                background: linear-gradient(180deg, rgba(45, 140, 255, 0.22), rgba(45, 140, 255, 0.12));
                color: var(--zoom-text-strong) !important;
                border: 1px solid rgba(76, 201, 255, 0.16);
            }

            [data-testid="stDataFrame"],
            [data-testid="stDataEditor"] {
                border: 1px solid var(--zoom-line);
                border-radius: 24px;
                overflow: hidden;
                box-shadow: 0 18px 40px rgba(1, 6, 20, 0.28);
                background: var(--zoom-panel);
            }

            .stExpander {
                border: 1px solid var(--zoom-line);
                border-radius: 22px;
                background: var(--zoom-panel);
            }

            .app-spacer-xl {
                height: 4.25rem;
            }

            .app-spacer-lg {
                height: 3rem;
            }

            .app-spacer-md {
                height: 1.85rem;
            }

            .stAlert {
                border-radius: 20px;
            }

            @media (max-width: 900px) {
                .status-grid {
                    grid-template-columns: 1fr 1fr;
                }
            }

            @media (max-width: 640px) {
                .status-grid {
                    grid-template-columns: 1fr;
                }
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_section_header(kicker: str, title: str, copy: str) -> None:
    """Render a consistent section heading."""
    st.markdown(
        f"""
        <div class="section-shell">
            <div class="section-kicker">{html.escape(kicker)}</div>
            <h2 class="section-title">{html.escape(title)}</h2>
            <p class="section-copy">{html.escape(copy)}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_spacer(size: str = "lg") -> None:
    """Render vertical breathing room between major sections."""
    if size not in {"md", "lg", "xl"}:
        size = "lg"
    st.markdown(f'<div class="app-spacer-{size}"></div>', unsafe_allow_html=True)


def build_data_uri_download_href(data: bytes, mime: str = "application/octet-stream") -> str:
    """Return a data URI that can be used inside an HTML download link."""
    encoded = base64.b64encode(data).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def render_metric_card(
    column: Any,
    label: str,
    value: Any,
    footnote: str = "",
    tone: str = "neutral",
) -> None:
    """Render a branded KPI card."""
    column.markdown(
        f"""
        <div class="metric-card metric-{html.escape(tone)}">
            <div class="metric-label">{html.escape(label)}</div>
            <div class="metric-value">{html.escape(str(value))}</div>
            <div class="metric-footnote">{html.escape(footnote)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_notice_card(title: str, copy: str, tone: str = "success") -> None:
    """Render a styled notice banner."""
    st.markdown(
        f"""
        <div class="notice-card notice-{html.escape(tone)}">
            <div class="notice-title">{html.escape(title)}</div>
            <p class="notice-copy">{html.escape(copy)}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_start_here_strip() -> None:
    """Render a clear first-step strip for instructors."""
    st.markdown(
        """
        <div class="status-strip">
            <div class="card-kicker" style="color: #4cc9ff; margin-bottom: 0.7rem;">Start here</div>
            <div class="status-grid">
                <div class="status-chip ok">
                    <strong>Required export 1</strong>
                    <span>Zoom transcript file in <code>.vtt</code> format.</span>
                </div>
                <div class="status-chip ok">
                    <strong>Required export 2</strong>
                    <span>Zoom participant report in <code>.csv</code> format.</span>
                </div>
                <div class="status-chip pending">
                    <strong>Optional but helpful</strong>
                    <span>Class roster and alias map for stronger identity matching.</span>
                </div>
                <div class="status-chip pending">
                    <strong>Final output</strong>
                    <span>One Excel workbook with Award, Do_Not_Award, Manual_Review, and audit sheets.</span>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_first_time_guide() -> None:
    """Render a concise guide for first-time instructors."""
    st.markdown(
        """
        <div class="guide-card">
            <div class="card-kicker" style="color: #4cc9ff;">Quick start for first-time users</div>
            <h3 class="card-title">Follow this sequence once and the workflow becomes straightforward.</h3>
            <ul class="guide-list">
                <li><span class="guide-step">1</span><span>Export the <strong>participant report CSV</strong> and <strong>transcript VTT</strong> from Zoom for the same meeting.</span></li>
                <li><span class="guide-step">2</span><span>Upload those two files first. Add a roster or alias map only if you already have them ready.</span></li>
                <li><span class="guide-step">3</span><span>Read the validation and matching sections before trusting final decisions, especially for display-name mismatches.</span></li>
                <li><span class="guide-step">4</span><span>Use the final decision table for overrides, then export the Excel workbook as the grading record.</span></li>
            </ul>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_upload_readiness(
    transcript_count: int,
    attendance_count: int,
    has_roster: bool,
    has_aliases: bool,
) -> None:
    """Render a simple readiness panel that tells the user what to do next."""
    transcript_ready = transcript_count > 0
    attendance_ready = attendance_count > 0
    ready_to_grade = transcript_ready and attendance_ready
    next_step = "Upload both the Zoom transcript VTT and participant CSV to unlock the grading workflow."
    if transcript_ready and not attendance_ready:
        next_step = "Upload the Zoom participant CSV next so attendance evidence can be verified."
    elif attendance_ready and not transcript_ready:
        next_step = "Upload the Zoom transcript VTT next so speaking evidence can be verified."
    elif ready_to_grade:
        next_step = "Core evidence is ready. Continue into validation, matching review, and final decision export."

    st.markdown(
        f"""
        <div class="status-strip">
            <div class="card-kicker" style="color: #4cc9ff; margin-bottom: 0.7rem;">Workspace readiness</div>
            <p class="section-copy" style="margin-bottom: 1rem; max-width: none;">{html.escape(next_step)}</p>
            <div class="status-grid">
                <div class="status-chip {'ok' if transcript_ready else 'pending'}">
                    <strong>Transcript</strong>
                    <span>{'Ready: one or more VTT files uploaded.' if transcript_ready else 'Still needed: upload at least one Zoom transcript VTT.'}</span>
                </div>
                <div class="status-chip {'ok' if attendance_ready else 'pending'}">
                    <strong>Attendance</strong>
                    <span>{'Ready: one or more participant CSV files uploaded.' if attendance_ready else 'Still needed: upload at least one Zoom participant CSV.'}</span>
                </div>
                <div class="status-chip {'ok' if has_roster else 'pending'}">
                    <strong>Roster</strong>
                    <span>{'Optional boost: roster uploaded for stronger matching.' if has_roster else 'Optional: upload a roster if students use inconsistent display names.'}</span>
                </div>
                <div class="status-chip {'ok' if ready_to_grade else 'pending'}">
                    <strong>Ready to grade</strong>
                    <span>{'Yes: core evidence is present and the grading workspace can run.' if ready_to_grade else 'Not yet: upload both a transcript and attendance export to unlock full grading output.'}</span>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if has_aliases:
        st.caption("Alias mapping is loaded, so known nickname and display-name variations can be matched more safely.")


def render_sidebar_config() -> AppConfig:
    """Render scoring and policy controls in the sidebar."""
    with st.sidebar:
        st.markdown(
            """
            <div class="saas-shell sidebar-panel">
                <div class="card-kicker">Control Tower</div>
                <h3>Grading policy</h3>
                <p>
                    Tune participation rules, identity safety thresholds, and weighted scoring before you review decisions.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        config = AppConfig()
        with st.expander("Participation policy", expanded=True):
            config.attendance_threshold_minutes = st.slider(
                "Attendance threshold (minutes)",
                min_value=1,
                max_value=180,
                value=DEFAULT_ATTENDANCE_THRESHOLD,
            )
            config.words_threshold = st.slider(
                "Speaking threshold (total words)",
                min_value=1,
                max_value=500,
                value=DEFAULT_WORD_THRESHOLD,
            )
            config.turns_threshold = st.slider(
                "Speaking threshold (turns)",
                min_value=1,
                max_value=20,
                value=DEFAULT_TURN_THRESHOLD,
            )
            config.span_threshold_enabled = st.toggle(
                "Enable speaking span threshold",
                value=False,
            )
            config.span_threshold_minutes = st.slider(
                "Speaking span threshold (minutes)",
                min_value=1,
                max_value=60,
                value=DEFAULT_SPAN_THRESHOLD,
                disabled=not config.span_threshold_enabled,
            )
            config.bonus_policy_mode = st.selectbox(
                "Bonus policy mode",
                options=BONUS_POLICY_OPTIONS,
                index=BONUS_POLICY_OPTIONS.index("attended and spoke"),
            )

        with st.expander("Identity matching safety", expanded=True):
            config.fuzzy_threshold = st.slider(
                "Fuzzy match threshold",
                min_value=0.50,
                max_value=1.00,
                value=DEFAULT_FUZZY_THRESHOLD,
                step=0.01,
                format="%.2f",
            )
            config.safe_auto_approval_threshold = st.slider(
                "Safe auto-approval threshold",
                min_value=0.50,
                max_value=1.00,
                value=DEFAULT_SAFE_AUTO_THRESHOLD,
                step=0.01,
                format="%.2f",
            )
            st.caption(
                "Anything below the safe threshold is pushed toward manual review instead of auto-crediting."
            )

        if config.bonus_policy_mode == "weighted score":
            with st.expander("Weighted score formula", expanded=True):
                st.caption("Formula = attendance component + word component + turn component")
                config.attendance_weight = st.slider(
                    "Attendance weight",
                    min_value=0.0,
                    max_value=2.0,
                    value=float(DEFAULT_ATTENDANCE_WEIGHT),
                    step=0.05,
                )
                config.word_weight = st.slider(
                    "Word weight",
                    min_value=0.0,
                    max_value=2.0,
                    value=float(DEFAULT_WORD_WEIGHT),
                    step=0.05,
                )
                config.turn_weight = st.slider(
                    "Turn weight",
                    min_value=0.0,
                    max_value=2.0,
                    value=float(DEFAULT_TURN_WEIGHT),
                    step=0.05,
                )
                config.attendance_cap_minutes = st.slider(
                    "Attendance cap (minutes)",
                    min_value=1,
                    max_value=240,
                    value=DEFAULT_ATTENDANCE_CAP,
                )
                config.word_cap = st.slider(
                    "Word cap",
                    min_value=1,
                    max_value=500,
                    value=DEFAULT_WORD_CAP,
                )
                config.turn_cap = st.slider(
                    "Turn cap",
                    min_value=1,
                    max_value=30,
                    value=DEFAULT_TURN_CAP,
                )
                config.weighted_threshold = st.slider(
                    "Weighted score threshold",
                    min_value=0.0,
                    max_value=3.0,
                    value=float(DEFAULT_WEIGHTED_THRESHOLD),
                    step=0.05,
                )
        return config


def render_filter_controls(config: AppConfig, meeting_options: Sequence[str]) -> AppConfig:
    """Render sidebar filters once meeting options are available."""
    with st.sidebar:
        st.markdown(
            """
            <div class="saas-shell sidebar-panel">
                <div class="card-kicker">Review Lens</div>
                <h3>Decision filters</h3>
                <p>
                    Narrow the workspace to the exact meeting, student, or review category you want to inspect.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        with st.expander("Workspace filters", expanded=True):
            config.combine_all_meetings = st.toggle("Combine all meetings view", value=True)
            if not config.combine_all_meetings:
                config.selected_meetings = tuple(
                    st.multiselect(
                        "Per-meeting filter",
                        options=list(meeting_options),
                        default=list(meeting_options[:1]),
                    )
                )
            config.search_text = st.text_input("Search student name", value="")
            config.show_only_award = st.toggle("Show only Award", value=False)
            config.show_only_do_not_award = st.toggle("Show only Do_Not_Award", value=False)
            config.show_only_manual_review = st.toggle("Show only Manual_Review", value=False)
            config.show_only_unmatched = st.toggle("Show only unmatched", value=False)
            config.show_only_low_confidence = st.toggle("Show only low-confidence", value=False)
    return config


def render_instructions() -> None:
    """Render the app title and high-level instructions."""
    render_start_here_strip()
    render_spacer("lg")
    hero_col, detail_col = st.columns([1.65, 1.0])
    with hero_col:
        st.markdown(
            """
            <div class="hero-shell">
                <div class="hero-kicker">Zoom-style grading workspace</div>
                <div class="hero-title">Review class participation with evidence, guardrails, and an export-ready audit trail.</div>
                <p class="hero-copy">
                    This workspace is built for instructors who need a dependable way to combine attendance records,
                    transcript speaking evidence, and manual review into one grading workflow.
                </p>
                <div class="pill-row">
                    <span class="pill">Attendance-backed decisions</span>
                    <span class="pill">Speaking evidence from transcripts</span>
                    <span class="pill">Low-confidence review routing</span>
                    <span class="pill">One-click Excel export</span>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with detail_col:
        st.markdown(
            """
            <div class="hero-panel">
                <div class="card-kicker" style="color: #4cc9ff;">What you need before you start</div>
                <h3>Minimal instructor upload pack</h3>
                <p>Bring the two core Zoom exports first. Everything else is optional support for cleaner matching and review.</p>
                <ul class="hero-list">
                    <li><span>Transcript</span><span>Zoom <code>.vtt</code> export</span></li>
                    <li><span>Attendance</span><span>Zoom participant <code>.csv</code></span></li>
                    <li><span>Roster</span><span>Optional <code>.csv</code> or <code>.xlsx</code></span></li>
                    <li><span>Aliases</span><span>Optional nickname mapping <code>.csv</code></span></li>
                </ul>
            </div>
            """,
            unsafe_allow_html=True,
        )

    render_spacer("xl")
    workflow_cols = st.columns(3)
    workflow_cards = [
        (
            "01",
            "Upload the Zoom exports",
            "Start with the transcript VTT and participant CSV for the meeting you want to grade.",
        ),
        (
            "02",
            "Check the risky cases",
            "Inspect validation warnings, low-confidence name matches, and rows routed into manual review.",
        ),
        (
            "03",
            "Export the workbook",
            "Download the structured Excel output with raw evidence, final decisions, and the audit log.",
        ),
    ]
    for column, (step, title, copy) in zip(workflow_cols, workflow_cards):
        column.markdown(
            f"""
            <div class="workflow-card">
                <div class="workflow-step">{html.escape(step)}</div>
                <h3 class="card-title">{html.escape(title)}</h3>
                <p class="card-copy">{html.escape(copy)}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    render_spacer("xl")
    render_first_time_guide()


def render_sample_downloads() -> None:
    """Render sample file download buttons."""
    render_section_header(
        "Sandbox pack",
        "Try the workflow with matching sample files",
        "Use the synthetic sample packet to preview the full upload, review, and export flow before running a live class batch.",
    )
    sample_files = generate_sample_files()
    sample_columns = st.columns(len(sample_files))
    for idx, (name, data) in enumerate(sample_files.items()):
        file_type = Path(name).suffix.replace(".", "").upper()
        download_href = build_data_uri_download_href(data)
        sample_columns[idx].markdown(
            f"""
            <div class="sample-card">
                <div class="sample-card-body">
                    <div class="card-kicker" style="color: #4cc9ff;">{html.escape(file_type)}</div>
                    <h3 class="sample-card-file">{html.escape(name)}</h3>
                    <p class="sample-card-copy">Download this sample file and upload it to explore the app end to end.</p>
                </div>
                <a class="sample-card-cta" href="{download_href}" download="{html.escape(name, quote=True)}">Download sample</a>
            </div>
            """,
            unsafe_allow_html=True,
        )
    render_spacer("xl")


def render_empty_workspace_state() -> None:
    """Render a polished onboarding state before evidence is uploaded."""
    render_section_header(
        "Getting started",
        "The app becomes a grading workspace as soon as core Zoom evidence is uploaded",
        "You only need two exports from Zoom to start generating decisions: a transcript VTT and a participant CSV.",
    )
    columns = st.columns(3)
    cards = [
        (
            "Required",
            "Export from Zoom",
            "Download the meeting transcript as a .vtt file and the participant report as a .csv file.",
        ),
        (
            "Optional",
            "Strengthen identity matching",
            "Add a roster file and an alias map if students frequently use nicknames or device names.",
        ),
        (
            "Output",
            "Receive a final workbook",
            "The app returns Award, Do_Not_Award, Manual_Review, and audit sheets in one Excel export.",
        ),
    ]
    for column, (kicker, title, copy) in zip(columns, cards):
        column.markdown(
            f"""
            <div class="info-card">
                <div class="card-kicker" style="color: #4cc9ff;">{html.escape(kicker)}</div>
                <h3 class="card-title">{html.escape(title)}</h3>
                <p class="card-copy">{html.escape(copy)}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_core_upload_notice(
    has_transcript: bool,
    has_attendance: bool,
) -> None:
    """Explain why full grading sections are hidden until core evidence is ready."""
    missing_parts: list[str] = []
    if not has_transcript:
        missing_parts.append("Zoom transcript VTT")
    if not has_attendance:
        missing_parts.append("Zoom participant CSV")
    missing_text = " and ".join(missing_parts) if missing_parts else "required Zoom files"

    render_section_header(
        "Next required step",
        "Complete the core Zoom upload pack to unlock grading decisions",
        "Raw previews can still be inspected below, but matching review, final decisions, and export stay hidden until both required Zoom evidence files are present.",
    )
    render_notice_card(
        "Core grading workflow is still locked.",
        f"Upload the missing {missing_text} next. Once both are present, the app will unlock identity matching, final review, and Excel export.",
        tone="warning",
    )


def render_validation_summary(issues: list[ValidationIssue]) -> None:
    """Render validation issues in a friendly summary."""
    render_section_header(
        "Quality checks",
        "Validation summary",
        "The app flags malformed files, duplicates, and missing evidence before those issues can leak into grading decisions.",
    )
    if not issues:
        render_notice_card(
            "No validation issues detected.",
            "Current uploads passed the app's parsing and deduplication checks.",
            tone="success",
        )
        return

    issue_df = pd.DataFrame([issue.__dict__ for issue in issues])
    error_count = int((issue_df["level"] == "error").sum())
    warning_count = int((issue_df["level"] == "warning").sum())
    stat_cols = st.columns(3)
    render_metric_card(stat_cols[0], "Errors", error_count, "Rows or files that need correction", "review")
    render_metric_card(stat_cols[1], "Warnings", warning_count, "Items worth checking before export", "warning")
    render_metric_card(stat_cols[2], "Total checks", len(issue_df), "Combined validation findings", "neutral")
    st.dataframe(issue_df, use_container_width=True, hide_index=True)


def render_metrics(
    raw_transcript_df: pd.DataFrame,
    raw_attendance_df: pd.DataFrame,
    final_df: pd.DataFrame,
) -> None:
    """Render top-line KPI cards."""
    render_section_header(
        "Decision room",
        "Current grading snapshot",
        "Use these live metrics to see how much evidence has been ingested and how the current policy is shaping final decisions.",
    )
    meeting_count = len(
        {
            value
            for value in pd.concat(
                [
                    raw_transcript_df.get("meeting_key", pd.Series(dtype=str)),
                    raw_attendance_df.get("meeting_key", pd.Series(dtype=str)),
                ],
                ignore_index=True,
            )
            if stringify(value)
        }
    )
    award_count = int((final_df.get("final_category", pd.Series(dtype=str)) == "Award").sum())
    do_not_award_count = int((final_df.get("final_category", pd.Series(dtype=str)) == "Do_Not_Award").sum())
    manual_review_count = int((final_df.get("final_category", pd.Series(dtype=str)) == "Manual_Review").sum())

    row_one = st.columns(4)
    render_metric_card(row_one[0], "Uploaded meetings", meeting_count, "Meetings represented in current evidence", "neutral")
    render_metric_card(row_one[1], "Transcript rows", len(raw_transcript_df), "Raw speaking cues parsed from VTT files", "neutral")
    render_metric_card(row_one[2], "Attendance rows", len(raw_attendance_df), "Participant records parsed from CSV exports", "neutral")
    render_metric_card(row_one[3], "Matched students", len(final_df), "Decision rows currently in the review workspace", "neutral")

    row_two = st.columns(3)
    render_metric_card(row_two[0], "Award", award_count, "Students recommended for bonus credit", "award")
    render_metric_card(row_two[1], "Do_Not_Award", do_not_award_count, "Students below the active policy threshold", "warning")
    render_metric_card(row_two[2], "Manual_Review", manual_review_count, "Rows that need instructor inspection", "review")


def render_upload_area() -> tuple[list[Any], list[Any], Any, Any]:
    """Render file upload widgets."""
    render_section_header(
        "Evidence workspace",
        "Upload the files for this grading batch",
        "The app accepts multiple transcripts and participant reports at once, then merges them into one instructor review workspace.",
    )
    guide_cols = st.columns(4)
    guide_cards = [
        ("Required", "Transcript", "Upload one or more Zoom transcript files in .vtt format."),
        ("Required", "Attendance", "Upload one or more Zoom participant exports in .csv format."),
        ("Optional", "Roster", "Add a roster in .csv or .xlsx to improve identity confidence."),
        ("Optional", "Aliases", "Add a nickname or device-name map in .csv format."),
    ]
    for column, (kicker, title, copy) in zip(guide_cols, guide_cards):
        column.markdown(
            f"""
            <div class="info-card">
                <div class="card-kicker" style="color: #4cc9ff;">{html.escape(kicker)}</div>
                <h3 class="card-title" style="font-size: 1.05rem;">{html.escape(title)}</h3>
                <p class="card-copy">{html.escape(copy)}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    render_spacer("md")
    upload_cols = st.columns([1.2, 1.0])
    with upload_cols[0]:
        st.markdown(
            """
            <div class="workspace-banner">
                <strong>Core Zoom evidence</strong><br />
                Upload the two required files from Zoom here. Attendance drives presence; transcripts drive speaking evidence.
            </div>
            """,
            unsafe_allow_html=True,
        )
        transcript_uploads = st.file_uploader(
            "Upload Zoom transcript files (.vtt)",
            type=["vtt"],
            accept_multiple_files=True,
            help="Export the meeting transcript from Zoom and upload one or more .vtt files.",
        )
        attendance_uploads = st.file_uploader(
            "Upload Zoom participant export files (.csv)",
            type=["csv"],
            accept_multiple_files=True,
            help="Upload one or more participant CSV exports from Zoom.",
        )

    with upload_cols[1]:
        st.markdown(
            """
            <div class="workspace-banner">
                <strong>Identity enrichment</strong><br />
                These files are optional, but they make the matching layer more reliable when display names are inconsistent.
            </div>
            """,
            unsafe_allow_html=True,
        )
        roster_upload = st.file_uploader(
            "Optional roster file (.csv or .xlsx)",
            type=["csv", "xlsx"],
            accept_multiple_files=False,
            help="Upload your class roster to improve canonical matching.",
        )
        alias_upload = st.file_uploader(
            "Optional alias mapping file (.csv)",
            type=["csv"],
            accept_multiple_files=False,
            help="Upload a two-column alias file such as alias_name and canonical_name.",
        )
    return transcript_uploads, attendance_uploads, roster_upload, alias_upload


def build_roster_selection_ui(raw_roster_table: pd.DataFrame) -> tuple[str, str]:
    """Render roster column selectors."""
    if raw_roster_table.empty:
        return "", ""
    columns = list(raw_roster_table.columns)
    detected = detect_columns(columns, ROSTER_HEADER_SYNONYMS)
    name_candidates = detect_column_candidates(columns, ROSTER_HEADER_SYNONYMS["name"])
    email_candidates = detect_column_candidates(columns, ROSTER_HEADER_SYNONYMS["email"])

    default_name = detected.get("name", columns[0])
    default_email = detected.get("email", "")

    with st.expander("Roster column mapping", expanded=len(name_candidates) != 1 or len(email_candidates) != 1):
        st.caption("Adjust these selectors if the detected roster columns are incorrect.")
        name_column = st.selectbox(
            "Roster name column",
            options=columns,
            index=columns.index(default_name),
            key="roster_name_column",
        )
        email_options = [""] + columns
        email_index = email_options.index(default_email) if default_email in email_options else 0
        email_column = st.selectbox(
            "Roster email column",
            options=email_options,
            index=email_index,
            key="roster_email_column",
            format_func=lambda value: value or "No email column",
        )
    return name_column, email_column


def render_previews(
    raw_transcript_df: pd.DataFrame,
    raw_attendance_df: pd.DataFrame,
    matched_df: pd.DataFrame,
    final_df: pd.DataFrame,
) -> None:
    """Render parsed data previews."""
    render_section_header(
        "Evidence ledger",
        "Parsed data previews",
        "Inspect the raw and processed tables before you finalize any grading decisions.",
    )
    tabs = st.tabs(
        [
            "Transcript",
            "Attendance",
            "Matches",
            "Decisions",
        ]
    )
    with tabs[0]:
        if raw_transcript_df.empty:
            render_notice_card(
                "No transcript rows available.",
                "Upload one or more VTT files to inspect parsed speaking evidence.",
                tone="warning",
            )
        else:
            preview = raw_transcript_df.drop(columns=["start_seconds", "end_seconds", "interval_key"], errors="ignore")
            st.caption(f"{len(preview)} transcript row(s) parsed.")
            st.dataframe(preview.head(MAX_PREVIEW_ROWS), use_container_width=True, hide_index=True)
    with tabs[1]:
        if raw_attendance_df.empty:
            render_notice_card(
                "No attendance rows available.",
                "Upload one or more participant CSV files to inspect attendance evidence.",
                tone="warning",
            )
        else:
            st.caption(f"{len(raw_attendance_df)} attendance row(s) parsed.")
            st.dataframe(raw_attendance_df.head(MAX_PREVIEW_ROWS), use_container_width=True, hide_index=True)
    with tabs[2]:
        if matched_df.empty:
            render_notice_card(
                "No matched evidence rows available.",
                "Once evidence is uploaded, the app will show its identity resolution output here.",
                tone="warning",
            )
        else:
            st.caption(f"{len(matched_df)} matched evidence row(s) in view.")
            st.dataframe(matched_df.head(MAX_PREVIEW_ROWS), use_container_width=True, hide_index=True)
    with tabs[3]:
        if final_df.empty:
            render_notice_card(
                "No final decision rows available.",
                "Final recommendation rows appear here after matching and policy scoring run.",
                tone="warning",
            )
        else:
            preview_columns = [
                "meeting_name",
                "canonical_student_name",
                "raw_attendance_name",
                "raw_transcript_speaker",
                "attended",
                "spoke",
                "match_method",
                "match_confidence",
                "final_category",
                "decision_reason",
            ]
            st.caption(f"{len(final_df)} decision row(s) currently in view.")
            st.dataframe(final_df[preview_columns].head(MAX_PREVIEW_ROWS), use_container_width=True, hide_index=True)


def render_matching_review(filtered_matched_df: pd.DataFrame) -> None:
    """Render a focused matching review table."""
    render_section_header(
        "Identity review",
        "Matching review section",
        "This workspace helps you audit how attendance names and transcript speakers were matched to canonical students.",
    )
    if filtered_matched_df.empty:
        render_notice_card(
            "No matching rows available for the current filters.",
            "Clear some filters or upload evidence to inspect the identity resolution layer.",
            tone="warning",
        )
        return

    stat_cols = st.columns(4)
    review_count = int(filtered_matched_df.get("identity_review_flag", pd.Series(dtype=bool)).fillna(False).sum())
    unmatched_count = int(
        filtered_matched_df.get("match_method", pd.Series(dtype=str)).astype(str).str.contains("unmatched", case=False, na=False).sum()
    )
    low_confidence_count = int(
        pd.to_numeric(filtered_matched_df.get("match_confidence", pd.Series(dtype=float)), errors="coerce").fillna(0).lt(DEFAULT_SAFE_AUTO_THRESHOLD).sum()
    )
    render_metric_card(stat_cols[0], "Rows in view", len(filtered_matched_df), "Evidence rows currently visible", "neutral")
    render_metric_card(stat_cols[1], "Needs review", review_count, "Identity rows flagged for instructor attention", "review")
    render_metric_card(stat_cols[2], "Unmatched", unmatched_count, "Rows with no safe canonical match", "warning")
    render_metric_card(stat_cols[3], "Low confidence", low_confidence_count, "Rows below the default auto-approval threshold", "warning")

    review_columns = [
        "meeting_name",
        "evidence_source",
        "raw_name",
        "canonical_student_name",
        "email",
        "match_method",
        "match_confidence",
        "matched_from_source",
        "identity_review_flag",
        "review_reason",
    ]
    available_columns = [column for column in review_columns if column in filtered_matched_df.columns]
    st.dataframe(
        filtered_matched_df[available_columns].head(MAX_PREVIEW_ROWS),
        use_container_width=True,
        hide_index=True,
    )


def render_final_review_editor(final_df: pd.DataFrame, config: AppConfig) -> pd.DataFrame:
    """Render the editable final review table and persist overrides."""
    render_section_header(
        "Decision console",
        "Final decision section",
        "This is the instructor handoff layer. Edit the approved fields, add notes, and override the final category when needed.",
    )
    if final_df.empty:
        render_notice_card(
            "No decision rows are available yet.",
            "Upload evidence to generate the final instructor review table.",
            tone="warning",
        )
        return final_df

    summary_cols = st.columns(3)
    award_count = int((final_df["final_category"] == "Award").sum())
    no_award_count = int((final_df["final_category"] == "Do_Not_Award").sum())
    manual_count = int((final_df["final_category"] == "Manual_Review").sum())
    render_metric_card(summary_cols[0], "Award in view", award_count, "Rows currently recommended for bonus credit", "award")
    render_metric_card(summary_cols[1], "Do not award", no_award_count, "Rows currently below the active policy", "warning")
    render_metric_card(summary_cols[2], "Manual review", manual_count, "Rows requiring instructor judgment", "review")
    st.caption(
        "Editable fields persist in session state during this browser session: canonical student name, attended, spoke, recommend award, manual review, notes, and reviewer override."
    )

    display_columns = [
        "meeting_name",
        "meeting_id",
        "canonical_student_name",
        "raw_attendance_name",
        "raw_transcript_speaker",
        "email",
        "attended",
        "spoke",
        "attendance_minutes",
        "speaking_turns",
        "speaking_words",
        "speaking_span_minutes",
        "join_count",
        "match_method",
        "match_confidence",
        "manual_review",
        "recommend_award",
        "reviewer_notes",
        "reviewer_override",
        "final_category",
        "decision_reason",
    ]
    editor_df = final_df.set_index("row_key")[display_columns].copy()
    edited_df = st.data_editor(
        editor_df,
        use_container_width=True,
        hide_index=True,
        key="final_review_editor",
        disabled=[column for column in display_columns if column not in EDITABLE_FIELDS],
        column_config={
            "canonical_student_name": st.column_config.TextColumn("Canonical student name"),
            "attended": st.column_config.CheckboxColumn("Attended"),
            "spoke": st.column_config.CheckboxColumn("Spoke"),
            "manual_review": st.column_config.CheckboxColumn("Manual review"),
            "recommend_award": st.column_config.CheckboxColumn("Recommend award"),
            "reviewer_notes": st.column_config.TextColumn("Reviewer notes"),
            "reviewer_override": st.column_config.SelectboxColumn(
                "Reviewer override",
                options=REVIEWER_OVERRIDE_OPTIONS,
            ),
            "match_confidence": st.column_config.NumberColumn("Match confidence", format="%.2f"),
            "attendance_minutes": st.column_config.NumberColumn("Attendance minutes", format="%.2f"),
            "speaking_span_minutes": st.column_config.NumberColumn("Speaking span minutes", format="%.2f"),
        },
    )

    if persist_editor_overrides(editor_df, edited_df):
        st.rerun()
    return apply_manual_overrides(final_df, config)


def render_export_section(
    config: AppConfig,
    raw_transcript_df: pd.DataFrame,
    raw_attendance_df: pd.DataFrame,
    speaker_agg_df: pd.DataFrame,
    attendance_agg_df: pd.DataFrame,
    matched_df: pd.DataFrame,
    final_df: pd.DataFrame,
    issues: list[ValidationIssue],
    uploaded_filenames: Sequence[str],
) -> None:
    """Render the Excel export section."""
    render_section_header(
        "Delivery",
        "Excel export section",
        "When the review workspace looks right, export the entire evidence trail and final decisions as one workbook.",
    )
    if final_df.empty and raw_transcript_df.empty and raw_attendance_df.empty:
        render_notice_card(
            "Upload data first to enable export.",
            "Once transcript or attendance evidence is available, the Excel workbook download will appear here.",
            tone="warning",
        )
        return

    workbook_bytes = build_excel_workbook(
        config=config,
        raw_transcript_df=raw_transcript_df,
        raw_attendance_df=raw_attendance_df,
        speaker_agg_df=speaker_agg_df,
        attendance_agg_df=attendance_agg_df,
        matched_df=matched_df,
        final_df=final_df,
        issues=issues,
        uploaded_filenames=uploaded_filenames,
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    export_cols = st.columns([1.15, 0.85])
    with export_cols[0]:
        st.markdown(
            """
            <div class="info-card">
                <div class="card-kicker" style="color: #4cc9ff;">Workbook contents</div>
                <h3 class="card-title">Everything needed for audit and grading</h3>
                <p class="card-copy">The export includes raw evidence, aggregated tables, decision sheets, and a timestamped audit log.</p>
                <ul class="sheet-list">
                    <li><span>Config</span><span>Policy settings and export timestamp</span></li>
                    <li><span>Raw evidence</span><span>Transcript and attendance imports</span></li>
                    <li><span>Aggregations</span><span>Speaker and attendance rollups</span></li>
                    <li><span>Decisions</span><span>Award, Do_Not_Award, Manual_Review</span></li>
                    <li><span>Audit log</span><span>Warnings, unmatched rows, low-confidence notes</span></li>
                </ul>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with export_cols[1]:
        st.markdown(
            f"""
            <div class="hero-panel" style="background: linear-gradient(180deg, rgba(33, 47, 89, 0.82), rgba(17, 25, 50, 0.98));">
                <div class="card-kicker" style="color: #4cc9ff;">Ready to ship</div>
                <h3>Export the grading workbook</h3>
                <p>{html.escape(str(len(final_df)))} decision row(s) and {html.escape(str(len(uploaded_filenames)))} uploaded file reference(s) will be bundled into one download.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.download_button(
            "Download Excel workbook",
            data=workbook_bytes,
            file_name=f"zoom_participation_grader_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# =========================
# Main app entrypoint
# =========================


def main() -> None:
    """Run the Streamlit app."""
    st.set_page_config(
        page_title=APP_TITLE,
        layout="wide",
        initial_sidebar_state="expanded",
    )
    render_global_styles()
    config = render_sidebar_config()
    render_instructions()
    render_sample_downloads()
    transcript_uploads, attendance_uploads, roster_upload, alias_upload = render_upload_area()
    render_spacer("md")
    render_upload_readiness(
        transcript_count=len(transcript_uploads or []),
        attendance_count=len(attendance_uploads or []),
        has_roster=roster_upload is not None,
        has_aliases=alias_upload is not None,
    )

    if not transcript_uploads and not attendance_uploads:
        render_spacer("xl")
        render_empty_workspace_state()
        return

    transcript_files, transcript_upload_issues = uploaded_to_memory(transcript_uploads or [], "transcript")
    attendance_files, attendance_upload_issues = uploaded_to_memory(attendance_uploads or [], "attendance")
    issues: list[ValidationIssue] = []
    issues.extend(transcript_upload_issues)
    issues.extend(attendance_upload_issues)

    raw_transcript_frames: list[pd.DataFrame] = []
    speaker_frames: list[pd.DataFrame] = []
    raw_attendance_frames: list[pd.DataFrame] = []

    for file in transcript_files:
        transcript_df, transcript_issues = parse_transcript_file(file["name"], file["data"])
        issues.extend(transcript_issues)
        raw_transcript_frames.append(transcript_df)

    raw_transcript_df = (
        pd.concat(raw_transcript_frames, ignore_index=True)
        if raw_transcript_frames
        else empty_df(TRANSCRIPT_COLUMNS + ["start_seconds", "end_seconds", "interval_key"])
    )
    speaker_agg_df = aggregate_transcript_speakers(raw_transcript_df)
    if not speaker_agg_df.empty:
        speaker_frames.append(speaker_agg_df)
    speaker_agg_df = (
        pd.concat(speaker_frames, ignore_index=True)
        if speaker_frames
        else empty_df(SPEAKER_AGG_COLUMNS + ["first_speaking_seconds", "last_speaking_seconds", "speaker_is_unknown", "source_file"])
    )

    for file in attendance_files:
        attendance_df, attendance_issues = parse_attendance_file(file["name"], file["data"])
        issues.extend(attendance_issues)
        raw_attendance_frames.append(attendance_df)

    raw_attendance_df = (
        pd.concat(raw_attendance_frames, ignore_index=True)
        if raw_attendance_frames
        else empty_df(ATTENDANCE_COLUMNS)
    )
    attendance_agg_df = aggregate_attendance(raw_attendance_df)

    raw_roster_table = empty_df([])
    roster_df = empty_df(ROSTER_COLUMNS)
    if roster_upload is not None:
        roster_data = roster_upload.getvalue()
        if roster_data:
            raw_roster_table, roster_issues = read_table_file(
                roster_upload.name, roster_data, ROSTER_HEADER_SYNONYMS
            )
            issues.extend(roster_issues)
            name_column, email_column = build_roster_selection_ui(raw_roster_table)
            roster_df, roster_prepare_issues = prepare_roster_df(
                raw_roster_table, name_column, email_column
            )
            issues.extend(roster_prepare_issues)
        else:
            issues.append(
                ValidationIssue("warning", "roster", roster_upload.name, "Empty roster file skipped.")
            )

    alias_df = empty_df(ALIAS_COLUMNS)
    if alias_upload is not None:
        alias_data = alias_upload.getvalue()
        if alias_data:
            raw_alias_table, alias_issues = read_table_file(
                alias_upload.name, alias_data, ALIAS_HEADER_SYNONYMS
            )
            issues.extend(alias_issues)
            alias_df, alias_prepare_issues = prepare_alias_df(raw_alias_table)
            issues.extend(alias_prepare_issues)
        else:
            issues.append(
                ValidationIssue("warning", "alias", alias_upload.name, "Empty alias file skipped.")
            )

    if raw_transcript_df.empty:
        render_notice_card(
            "No transcript uploaded.",
            "Speaking evidence will be unavailable until one or more VTT transcript files are added.",
            tone="warning",
        )
    if raw_attendance_df.empty:
        render_notice_card(
            "No attendance uploaded.",
            "Transcript-only rows will remain manual review and cannot auto-award until a participant CSV is uploaded.",
            tone="warning",
        )

    core_inputs_ready = not raw_transcript_df.empty and not raw_attendance_df.empty
    if not core_inputs_ready:
        empty_final = empty_df(
            FINAL_COLUMNS
            + [
                "meeting_key",
                "weighted_score",
                "final_category",
                "has_attendance_record",
                "has_transcript_record",
                "spoke_strong",
                "matched_from_source",
                "student_merge_key",
                "row_key",
                "unmatched",
                "low_confidence",
                "manual_review_reason",
                "identity_confidence_note",
                "source_file",
            ]
        )
        render_spacer("xl")
        render_validation_summary(issues)
        render_spacer("xl")
        render_metrics(raw_transcript_df, raw_attendance_df, empty_final)
        render_spacer("xl")
        render_previews(raw_transcript_df, raw_attendance_df, empty_df(MATCHED_COLUMNS), empty_final)
        render_spacer("xl")
        render_core_upload_notice(
            has_transcript=not raw_transcript_df.empty,
            has_attendance=not raw_attendance_df.empty,
        )
        return

    global_candidates, meeting_candidates = build_candidates(roster_df, attendance_agg_df)
    alias_lookup = build_alias_lookup(alias_df)
    attendance_matched_df = match_attendance_records(
        attendance_agg_df,
        global_candidates,
        meeting_candidates,
        alias_lookup,
        config,
    )
    transcript_matched_df = match_transcript_records(
        speaker_agg_df,
        global_candidates,
        meeting_candidates,
        alias_lookup,
        config,
    )
    matched_df = pd.concat(
        [attendance_matched_df, transcript_matched_df],
        ignore_index=True,
    ) if not attendance_matched_df.empty or not transcript_matched_df.empty else empty_df(MATCHED_COLUMNS)

    final_df = build_final_decision_table(
        matched_df=matched_df,
        raw_transcript_df=raw_transcript_df,
        roster_present=not roster_df.empty,
        config=config,
    )
    final_df = apply_manual_overrides(final_df, config)

    meeting_options = sorted(
        {
            stringify(value)
            for value in pd.concat(
                [
                    raw_transcript_df.get("meeting_name", pd.Series(dtype=str)),
                    raw_attendance_df.get("meeting_name", pd.Series(dtype=str)),
                ],
                ignore_index=True,
            )
            if stringify(value)
        }
    )
    config = render_filter_controls(config, meeting_options)

    render_spacer("xl")
    render_validation_summary(issues)
    render_spacer("xl")
    render_metrics(raw_transcript_df, raw_attendance_df, final_df)

    filtered_matched_df = apply_filters(
        matched_df.merge(
            final_df[["meeting_key", "student_merge_key", "final_category", "unmatched", "low_confidence"]],
            on=["meeting_key", "student_merge_key"],
            how="left",
        ) if not matched_df.empty and "student_merge_key" in matched_df.columns else matched_df,
        config,
    )
    filtered_final_df = apply_filters(final_df, config)

    render_spacer("xl")
    render_previews(raw_transcript_df, raw_attendance_df, filtered_matched_df, filtered_final_df)
    render_spacer("xl")
    render_matching_review(filtered_matched_df)
    render_spacer("xl")
    reviewed_filtered_df = render_final_review_editor(filtered_final_df, config)

    if not reviewed_filtered_df.empty:
        full_final_df = apply_manual_overrides(final_df, config)
    else:
        full_final_df = final_df

    uploaded_filenames = [file["name"] for file in transcript_files + attendance_files]
    if roster_upload is not None:
        uploaded_filenames.append(roster_upload.name)
    if alias_upload is not None:
        uploaded_filenames.append(alias_upload.name)

    render_spacer("xl")
    render_export_section(
        config=config,
        raw_transcript_df=raw_transcript_df,
        raw_attendance_df=raw_attendance_df,
        speaker_agg_df=speaker_agg_df,
        attendance_agg_df=attendance_agg_df,
        matched_df=matched_df,
        final_df=full_final_df,
        issues=issues,
        uploaded_filenames=uploaded_filenames,
    )


if __name__ == "__main__":
    main()
